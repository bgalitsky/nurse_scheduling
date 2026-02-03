#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Streamlit app: physician shift scheduling (CSP-like heuristic) + XLSX export.

Features:
- Editable inputs (doctors + FTE, priorities, vacations, cabins, holidays)
- Recompute schedule button
- On-screen table (doctor x days)
- Download XLSX (color-coded + 6 sheets)

Deps:
  pip install streamlit openpyxl pandas

Run:
  streamlit run doct_scheduler_app.py
"""

import re
import random
import datetime as dt
from dataclasses import dataclass, field
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Set, Optional
from io import BytesIO

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# Data model
# ----------------------------

@dataclass
class Doctor:
    name: str
    fte: float = 1.0  # 1.0 or 0.5
    priorities: List[str] = field(default_factory=list)

@dataclass
class Vacation:
    name: str
    start: dt.date
    end: dt.date  # inclusive


# ----------------------------
# Parsing helpersuse_container_width
# ----------------------------

def parse_date(s: str, year: int, month: int) -> dt.date:
    s = s.strip()
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        d, m, y = map(int, s.split("."))
        return dt.date(y, m, d)
    if re.match(r"^\d{2}\.\d{2}$", s):
        d, m = map(int, s.split("."))
        return dt.date(year, m, d)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        y, m, d = map(int, s.split("-"))
        return dt.date(y, m, d)
    raise ValueError(f"Bad date: {s}")

def daterange(a: dt.date, b: dt.date):
    cur = a
    while cur <= b:
        yield cur
        cur += dt.timedelta(days=1)

def is_weekend(d: dt.date) -> bool:
    return d.weekday() >= 5

def all_days_in_month(year: int, month: int) -> List[dt.date]:
    days = []
    d = dt.date(year, month, 1)
    while d.month == month:
        days.append(d)
        d += dt.timedelta(days=1)
    return days

def working_days_in_month(year: int, month: int) -> List[dt.date]:
    return [d for d in all_days_in_month(year, month) if not is_weekend(d)]

def build_vac_map(vacs: List["Vacation"]) -> Dict[str, Set[dt.date]]:
    m = defaultdict(set)
    for v in vacs:
        for day in daterange(v.start, v.end):
            m[v.name].add(day)
    return m

def parse_doctors_csv(text: str) -> List["Doctor"]:
    out: List[Doctor] = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if not parts or not parts[0]:
            continue
        name = parts[0]
        fte = float(parts[1]) if len(parts) >= 2 and parts[1] else 1.0
        out.append(Doctor(name=name, fte=fte))
    return out

def parse_priorities_csv(text: str) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        if len(parts) < 2:
            continue
        name = parts[0]
        cabs = [c.strip() for c in parts[1].split("|") if c.strip()]
        out[name] = cabs
    return out

def parse_vacations_csv(text: str, year: int, month: int) -> List["Vacation"]:
    out: List[Vacation] = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        name = parts[0]
        start = parse_date(parts[1], year, month)
        end = parse_date(parts[2], year, month)
        out.append(Vacation(name=name, start=start, end=end))
    return out

def parse_cabins(text: str) -> List[str]:
    items = re.split(r"[,\s]+", text.strip())
    return [i for i in (x.strip() for x in items) if i]

def parse_holidays(text: str, year: int, month: int) -> Set[dt.date]:
    out=set()
    if not text.strip():
        return out
    for tok in re.split(r"[,\s]+", text.strip()):
        if not tok.strip():
            continue
        out.add(parse_date(tok.strip(), year, month))
    return out


# ----------------------------
# CSP-like scheduler (heuristic with hard constraints)
# ----------------------------

def calc_required_shifts(doc: Doctor, workdays: List[dt.date], vac_days: Set[dt.date]) -> int:
    base = 23 if doc.fte >= 0.99 else 11
    vac_wd = sum(1 for d in workdays if d in vac_days)
    return max(0, base - vac_wd)

def score_candidate(doc: Doctor, cabin: str, shift_code: str,
                    remaining_need: int, streak: int,
                    used_today: bool, on_vac: bool,
                    weekend_total: int) -> float:
    if used_today or on_vac:
        return -1e9
    if streak >= 6:
        return -1e9
    s = 0.0
    s += remaining_need * 10.0
    s -= max(0, -remaining_need) * 20.0
    if cabin in doc.priorities:
        s += 8.0
    if doc.fte < 1.0:
        s -= 3.0
    if shift_code == "р":
        s -= weekend_total * 1.0
    return s

def schedule_month(doctors: List[Doctor],
                   vacations: List[Vacation],
                   cabins: List[str],
                   year: int,
                   month: int,
                   holidays: Optional[Set[dt.date]] = None,
                   seed: int = 42,
                   free_label: str = "свободно"):

    random.seed(seed)
    days = all_days_in_month(year, month)
    workdays = working_days_in_month(year, month)
    holidays = holidays or set()
    vac_map = build_vac_map(vacations)

    required = {doc.name: calc_required_shifts(doc, workdays, vac_map.get(doc.name, set())) for doc in doctors}

    total = Counter()
    weekend_total = Counter()
    streak = Counter()

    sched: Dict[str, Dict[dt.date, Tuple[str, str]]] = {d.name: {} for d in doctors}
    for doc in doctors:
        for day in days:
            if day in vac_map.get(doc.name, set()):
                sched[doc.name][day] = ("от", "")

    slot_assign: Dict[dt.date, List[Tuple[str, str, str]]] = {}

    cab_pop = Counter()
    for doc in doctors:
        for c in doc.priorities:
            cab_pop[c] += 1

    for day in days:
        wkend = is_weekend(day) or (day in holidays)
        shifts = ["р"] if wkend else ["у", "в"]
        slots = [(sh, cab) for sh in shifts for cab in cabins]
        slots.sort(key=lambda x: (-cab_pop[x[1]], x[0], x[1]))

        used_today: Set[str] = set()
        day_assign: List[Tuple[str, str, str]] = []

        for sh, cab in slots:
            best_name = None
            best_score = -1e18
            order = doctors[:]
            random.shuffle(order)

            for doc in order:
                name = doc.name
                on_vac = day in vac_map.get(name, set())
                rem = required[name] - total[name]
                sc = score_candidate(doc, cab, sh, rem, streak[name], (name in used_today), on_vac, weekend_total[name])
                if sc > best_score:
                    best_score = sc
                    best_name = name

            if best_name is None or best_score < -1e8:
                day_assign.append((sh, cab, free_label))
            else:
                day_assign.append((sh, cab, best_name))
                used_today.add(best_name)

        slot_assign[day] = day_assign

        worked_today = {}
        for sh, cab, who in day_assign:
            if who != free_label:
                worked_today[who] = (sh, cab)

        for doc in doctors:
            name = doc.name
            if day in vac_map.get(name, set()):
                streak[name] = 0
                continue
            if name in worked_today:
                sh, cab = worked_today[name]
                sched[name][day] = (sh, cab)
                total[name] += 1
                if sh == "р":
                    weekend_total[name] += 1
                streak[name] += 1
            else:
                sched[name][day] = ("-", "")
                streak[name] = 0

    deviation = {name: sum(1 for d in days if sched[name][d][0] in ("у", "в", "р")) - required[name] for name in required}
    return days, required, sched, slot_assign, deviation


# ----------------------------
# Excel export (to bytes)
# ----------------------------

def export_xlsx_bytes(doctors: List[Doctor],
                      days: List[dt.date],
                      sched: Dict[str, Dict[dt.date, Tuple[str, str]]],
                      required: Dict[str, int],
                      slot_assign: Dict[dt.date, List[Tuple[str, str, str]]],
                      cabins: List[str],
                      free_label: str = "свободно") -> bytes:

    wb = Workbook()
    ws = wb.active
    ws.title = "График врачей"

    fill_morning = PatternFill("solid", fgColor="A7D8FF")
    fill_evening = PatternFill("solid", fgColor="FFB6C1")
    fill_weekend = PatternFill("solid", fgColor="A7F3A7")
    fill_vac = PatternFill("solid", fgColor="C0C0C0")
    fill_off = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Врач").font = header_font
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=d.day)
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 11

    last_col = 2 + len(days)
    ws.cell(row=1, column=last_col, value="Смен").font = header_font
    ws.cell(row=1, column=last_col + 1, value="Комментарий").font = header_font

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(last_col)].width = 8
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 24

    for r, doc in enumerate(doctors, start=2):
        ws.cell(row=r, column=1, value=doc.name).border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)

        shift_count = 0
        for i, day in enumerate(days, start=2):
            code, cab = sched[doc.name][day]
            txt = code
            if code in ("у", "в", "р"):
                txt = f"{code} ({cab})"
                shift_count += 1

            cell = ws.cell(row=r, column=i, value=txt)
            cell.alignment = center
            cell.border = border

            if code == "у":
                cell.fill = fill_morning
            elif code == "в":
                cell.fill = fill_evening
            elif code == "р":
                cell.fill = fill_weekend
            elif code == "от":
                cell.fill = fill_vac
            else:
                cell.fill = fill_off

        ws.cell(row=r, column=last_col, value=shift_count).alignment = center
        ws.cell(row=r, column=last_col).border = border

        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        ws.cell(row=r, column=last_col + 1, value=f"ставка {rate}, норма {required[doc.name]}").border = border
        ws.cell(row=r, column=last_col + 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("Сводка по врачам")
    ws2.append(["Врач", "Ставка", "Норма смен", "Факт смен", "Отклонение", "Приоритеты"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        ws2.append([doc.name, doc.fte, required[doc.name], fact, fact - required[doc.name], ", ".join(doc.priorities)])

    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=6):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["F"].width = 26

    # Sheet 3
    ws3 = wb.create_sheet("Загрузка кабинетов")
    ws3.append(["Дата", "День", "Смена", "Кабинет", "Врач"])
    for c in ws3[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for d in days:
        for sh, cab, who in slot_assign[d]:
            ws3.append([d.isoformat(), day_name[d.weekday()], sh, cab, who])

    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 6
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 34

    # Sheet 4
    ws4 = wb.create_sheet("Общая статистика")
    ws4.append(["Показатель", "Значение"])
    ws4["A1"].font = header_font
    ws4["B1"].font = header_font

    total_slots = sum(len(slot_assign[d]) for d in days)
    free_slots = sum(1 for d in days for sh, cab, who in slot_assign[d] if who == free_label)
    ws4.append(["Месяц", f"{days[0].strftime('%B')} {days[0].year}"])
    ws4.append(["Кабинетов", len(cabins)])
    ws4.append(["Всего слотов (кабинет-смена)", total_slots])
    ws4.append(["Свободно (не заполнено)", free_slots])
    ws4.append(["Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0])

    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18

    # Sheet 5
    ws5 = wb.create_sheet("Обоснование отклонений")
    ws5.append(["Врач", "Норма", "Факт", "Отклонение", "Пояснение"])
    for c in ws5[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        dev = fact - required[doc.name]
        vac_count = sum(1 for d in days if sched[doc.name][d][0] == "от")
        expl = []
        if vac_count:
            expl.append(f"отпуск: {vac_count} дн.")
        if dev > 0:
            expl.append(f"доп. смены: {dev}")
        elif dev < 0:
            expl.append(f"не добрали: {-dev}")
        else:
            expl.append("норма выполнена")
        ws5.append([doc.name, required[doc.name], fact, dev, "; ".join(expl)])

    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["E"].width = 38

    # Sheet 6
    ws6 = wb.create_sheet("Кабинетов на врача")
    ws6.append(["Врач", "Кабинет", "Кол-во смен"])
    for c in ws6[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        cab_counter = Counter()
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                cab_counter[cab] += 1
        for cab, cnt in cab_counter.most_common():
            ws6.append([doc.name, cab, cnt])

    for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws6.column_dimensions["A"].width = 34
    ws6.column_dimensions["B"].width = 10
    ws6.column_dimensions["C"].width = 14

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------

st.set_page_config(page_title="График смен врачей (CSP)", layout="wide")
st.title("График смен врачей — CSP")

with st.sidebar:
    st.header("Параметры")
    year = st.number_input("Год", min_value=2020, max_value=2035, value=2025, step=1)
    month = st.number_input("Месяц", min_value=1, max_value=12, value=10, step=1)
    seed = st.number_input("Seed (воспроизводимость)", min_value=0, max_value=10_000, value=42, step=1)

    st.caption("Форматы: CSV строки. Комментарии начинаются с #.")

    default_doctors = """\
Боброва Эльвира Анатольевна,1
Буданова Татьяна Владимировна,0.5
Варвус Иван Михайлович,1
Гагкаева Мария Аркадьевна,1
Гущина Юлия Викторовна,1
Джафаров Башир Темирханович,0.5
Догузова Александра Гочевна,1
Евдокимова Жанна Владимировна,1
Егуренкова Виктория Сергеевна,1
Захаров Александр Сергеевич,1
Исаев Абсалам Темурланович,1
Китков Игорь Игоревич,1
Кокорина Серафима Евгеньевна,1
Королькова Александра Викторовна,0.5
Короткова Полина Александровна,1
Лукьяненко Владимир Александрович,1
Маркин Александр Андреевич,1
Марфутов Василий Васильевич,1
Медведева Мария Вячеславовна,1
Найданова Лидия Вадимовна,1
Нам Ирина Николаевна,1
Петрова Полина Юрьевна,1
Пономарева Юлия Константиновна,1
Прокофьев Евгений Андреевич,1
Пусь Юлия Владимировна,1
Танкиева Фатима Умат-Гиреевна,1
Тетерев Ярослав Тарасович,1
Федорова Оксана Сергеевна,1
Харитонова Наталья Игоревна,1
Шепель Артем Олегович,1
Юдин Алексей Александрович,1
Юдина Виктория Дмитриевна,1
"""
    default_priorities = """\
Пономарева Юлия Константиновна,2А03
Лукьяненко Владимир Александрович,2А03
Королькова Александра Викторовна,2А03
Харитонова Наталья Игоревна,2А04
Исаев Абсалам Темурланович,2А04
Короткова Полина Александровна,2А04
Гущина Юлия Викторовна,2А05
Федорова Оксана Сергеевна,2А05
Кокорина Серафима Евгеньевна,2А05
Егуренкова Виктория Сергеевна,2А06
Евдокимова Жанна Владимировна,2А06
Буданова Татьяна Владимировна,2А06
Боброва Эльвира Анатольевна,2А07
Марфутов Василий Васильевич,2А07
Догузова Александра Гочевна,2А07
Медведева Мария Вячеславовна,2А08
Юдин Алексей Александрович,2А08
Юдина Виктория Дмитриевна,2А33
Шепель Артем Олегович,2А33
Китков Игорь Игоревич,2А34
Гагкаева Мария Аркадьевна,2А34
Варвус Иван Михайлович,2А49
Маркин Александр Андреевич,2А49
Тетерев Ярослав Тарасович,2А52
Джафаров Башир Темирханович,2А52
Петрова Полина Юрьевна,2А52
Пусь Юлия Владимировна,2А54
Прокофьев Евгений Андреевич,2А54
Захаров Александр Сергеевич,2А54
Найданова Лидия Вадимовна,2А55
Нам Ирина Николаевна,2А55|2А54
Танкиева Фатима Умат-Гиреевна,2А55
"""
    default_vac = """\
Боброва Эльвира Анатольевна,01.10.2025,12.10.2025
Тетерев Ярослав Тарасович,29.09.2025,12.10.2025
Исаев Абсалам Темурланович,06.10.2025,19.10.2025
Гагкаева Мария Аркадьевна,01.10.2025,14.10.2025
Гагкаева Мария Аркадьевна,20.10.2025,20.10.2025
Гагкаева Мария Аркадьевна,27.10.2025,27.10.2025
Егуренкова Виктория Сергеевна,06.10.2025,19.10.2025
"""
    default_cabins = "2А03, 2А04, 2А05, 2А06, 2А07, 2А08, 2А33, 2А34, 2А49, 2А52, 2А54, 2А55"
    default_holidays = ""

    doctors_text = st.text_area("Врачи (name,fte)", value=default_doctors, height=240)
    prio_text = st.text_area("Приоритеты (name,cab1|cab2)", value=default_priorities, height=240)
    vac_text = st.text_area("Отпуска (name,start,end)", value=default_vac, height=160)
    cabins_text = st.text_area("Кабинеты", value=default_cabins, height=70)
    holidays_text = st.text_area("Праздники/доп. выходные (опц.) dd.mm.yyyy", value=default_holidays, height=60)

    recompute_clicked = st.button("Recompute", type="primary", width='stretch')

if "computed" not in st.session_state:
    st.session_state.computed = False

def compute():
    doctors = parse_doctors_csv(doctors_text)
    pr_map = parse_priorities_csv(prio_text)
    for d in doctors:
        d.priorities = pr_map.get(d.name, [])
    vacations = parse_vacations_csv(vac_text, int(year), int(month))
    cabins = parse_cabins(cabins_text)
    holidays = parse_holidays(holidays_text, int(year), int(month))

    if not doctors:
        raise ValueError("Список врачей пуст.")
    if not cabins:
        raise ValueError("Список кабинетов пуст.")

    days, required, sched, slot_assign, dev = schedule_month(
        doctors=doctors,
        vacations=vacations,
        cabins=cabins,
        year=int(year),
        month=int(month),
        holidays=holidays,
        seed=int(seed),
    )
    st.session_state.days = days
    st.session_state.doctors = doctors
    st.session_state.vacations = vacations
    st.session_state.cabins = cabins
    st.session_state.holidays = holidays
    st.session_state.required = required
    st.session_state.sched = sched
    st.session_state.slot_assign = slot_assign
    st.session_state.dev = dev
    st.session_state.computed = True

if recompute_clicked or not st.session_state.computed:
    try:
        with st.spinner("Считаю расписание..."):
            compute()
        st.success("Готово.")
    except Exception as e:
        st.session_state.computed = False
        st.error(f"Ошибка: {e}")

if st.session_state.computed:
    days: List[dt.date] = st.session_state.days
    doctors: List[Doctor] = st.session_state.doctors
    required = st.session_state.required
    sched = st.session_state.sched
    slot_assign = st.session_state.slot_assign
    cabins = st.session_state.cabins

    st.subheader("Лист 1 — График врачей")
    rows = []
    for doc in doctors:
        row = {"Врач": doc.name}
        shift_count = 0
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                row[str(d.day)] = f"{code} ({cab})"
                shift_count += 1
            else:
                row[str(d.day)] = code
        row["Смен"] = shift_count
        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        row["Комментарий"] = f"ставка {rate}, норма {required[doc.name]}"
        rows.append(row)

    df = pd.DataFrame(rows)
    st.dataframe(df, width='stretch', height=520)

    st.subheader("Сводка по врачам")
    summ_rows = []
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у","в","р"))
        summ_rows.append({
            "Врач": doc.name,
            "Ставка": doc.fte,
            "Норма смен": required[doc.name],
            "Факт смен": fact,
            "Отклонение": fact - required[doc.name],
            "Приоритеты": ", ".join(doc.priorities),
        })
    st.dataframe(pd.DataFrame(summ_rows), width='stretch', height=320)

    xlsx_bytes = export_xlsx_bytes(
        doctors=doctors,
        days=days,
        sched=sched,
        required=required,
        slot_assign=slot_assign,
        cabins=cabins,
    )
    filename = f"график_{int(year)}_{int(month):02d}.xlsx"
    st.download_button(
        label="Скачать XLSX (6 листов, цветной)",
        data=xlsx_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width='stretch',
    )

    with st.expander("Загрузка кабинетов (детально)"):
        load_rows = []
        day_name = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
        for d in days:
            for sh, cab, who in slot_assign[d]:
                load_rows.append({
                    "Дата": d.isoformat(),
                    "День": day_name[d.weekday()],
                    "Смена": sh,
                    "Кабинет": cab,
                    "Врач": who,
                })
        st.dataframe(pd.DataFrame(load_rows), width='stretch', height=420)
