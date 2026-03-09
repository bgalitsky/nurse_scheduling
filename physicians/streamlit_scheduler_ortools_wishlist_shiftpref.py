#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Streamlit app: physician shift scheduling with OR-Tools CP-SAT + XLSX export.

Implements the user's critical constraints:
- Weekdays: 2 shifts (утро 7:30–13:30, вечер 14:20–20:20)
- Weekends/holidays: 1 shift (р 9:00–18:00)
- One doctor: max 1 shift/day
- Max 5 consecutive working days (no 6-in-a-row)
- Vacations: no shifts assigned
- Norm per month (working days minus vacation working days):
    Full-time: 22 shifts
    Half-time: 11 shifts
- Extra shifts: only for doctors who agreed (optional input list). Others cannot exceed norm.
- Priorities:
    * Half-time doctors get lower priority weight.
    * If cabin has >=3 doctors who list it as their top priority, one doctor is selected (deterministically)
      and loses that priority (so they have no priority cabin).
    * If doctor is on vacation, they are not assigned anyway.

Output:
- On-screen schedule table (doctor x days) similar to the example PDF layout fileciteturn0file0
- Downloadable XLSX with 6 sheets and color coding.

Deps:
  pip install streamlit pandas openpyxl ortools

Run:
  streamlit run streamlit_scheduler_ortools.py
"""

import re
import datetime as dt
from dataclasses import dataclass, field
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Set, Optional
from io import BytesIO

import pandas as pd
import streamlit as st
from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# Data model
# ----------------------------

@dataclass
class Doctor:
    name: str
    fte: float = 1.0
    priorities: List[str] = field(default_factory=list)

@dataclass
class Vacation:
    name: str
    start: dt.date
    end: dt.date  # inclusive


# ----------------------------
# Helpers
# ----------------------------

def parse_date(s: str, year: int, month: int) -> dt.date:
    s = s.strip()
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        d, m, y = map(int, s.split("."))
        return dt.date(y, m, d)
    if re.match(r"^\d{2}\.\d{2}$", s):
        d, m = map(int, s.split("."))
        return dt.date(year, m, d)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        y, m, d = map(int, s.split("-"))
        return dt.date(y, m, d)
    raise ValueError(f"Bad date: {s}")

def daterange(a: dt.date, b: dt.date):
    cur = a
    while cur <= b:
        yield cur
        cur += dt.timedelta(days=1)

def is_weekend(d: dt.date) -> bool:
    return d.weekday() >= 5

def all_days_in_month(year: int, month: int) -> List[dt.date]:
    days = []
    d = dt.date(year, month, 1)
    while d.month == month:
        days.append(d)
        d += dt.timedelta(days=1)
    return days

def working_days_in_month(year: int, month: int) -> List[dt.date]:
    return [d for d in all_days_in_month(year, month) if not is_weekend(d)]

def parse_doctors_csv(text: str) -> List[Doctor]:
    """
    CSV lines: name, fte(1 or 0.5)
    """
    out: List[Doctor] = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if not parts or not parts[0]:
            continue
        name = parts[0]
        fte = float(parts[1]) if len(parts) >= 2 and parts[1] else 1.0
        out.append(Doctor(name=name, fte=fte))
    return out

def parse_priorities_csv(text: str) -> Dict[str, List[str]]:
    """
    CSV lines: name, cabin1|cabin2|...
    """
    out: Dict[str, List[str]] = {}
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        if len(parts) < 2:
            continue
        name = parts[0]
        cabs = [c.strip() for c in parts[1].split("|") if c.strip()]
        out[name] = cabs
    return out

def parse_vacations_csv(text: str, year: int, month: int) -> List[Vacation]:
    out: List[Vacation] = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        name = parts[0]
        start = parse_date(parts[1], year, month)
        end = parse_date(parts[2], year, month)
        out.append(Vacation(name=name, start=start, end=end))
    return out

def parse_cabins(text: str) -> List[str]:
    items = re.split(r"[,\s]+", text.strip())
    return [i for i in (x.strip() for x in items) if i]

def parse_holidays(text: str, year: int, month: int) -> Set[dt.date]:
    out=set()
    if not text.strip():
        return out
    for tok in re.split(r"[,\s]+", text.strip()):
        if not tok.strip():
            continue
        out.add(parse_date(tok.strip(), year, month))
    return out

def parse_yes_list(text: str) -> Set[str]:
    """One name per line (or CSV first column)."""
    out=set()
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        out.add(line.split(",")[0].strip())
    return out


def parse_shift_pref_csv(text: str) -> Dict[str, Optional[str]]:
    """CSV lines: name, pref_shift where pref_shift in {'у','в','нет','none',''}.
    Returns dict name -> 'у'/'в'/None.
    """
    out: Dict[str, Optional[str]] = {}
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        name = parts[0]
        pref = parts[1].strip().lower() if len(parts) > 1 else ""
        pref = _norm_cyr(pref) if '_norm_cyr' in globals() else pref
        if pref in ("у", "утро", "morning"):
            out[name] = "у"
        elif pref in ("в", "вечер", "evening"):
            out[name] = "в"
        elif pref in ("нет", "none", "-", ""):
            out[name] = None
        else:
            raise ValueError(f"Неверное предпочтение смены для '{name}': '{parts[1] if len(parts)>1 else ''}'. Используйте у/в/нет")
    return out

def build_vac_map(vacs: List[Vacation]) -> Dict[str, Set[dt.date]]:
    m = defaultdict(set)
    for v in vacs:
        for day in daterange(v.start, v.end):
            m[v.name].add(day)
    return m



# ----------------------------
# Wish-list parsing (optional, from XLSX)
# ----------------------------

_LATIN_TO_CYR = str.maketrans({
    "A":"А","B":"В","C":"С","E":"Е","H":"Н","K":"К","M":"М","O":"О","P":"Р","T":"Т","X":"Х","Y":"У",
    "a":"а","b":"в","c":"с","e":"е","h":"н","k":"к","m":"м","o":"о","p":"р","t":"т","x":"х","y":"у",
})

def _norm_cyr(s: str) -> str:
    return (s or "").translate(_LATIN_TO_CYR).replace("Ё","Е").replace("ё","е").strip()

def doctor_key_fullname(full_name: str) -> str:
    """'Иванов Иван Петрович' -> 'ИВАНОВ И.П.'"""
    parts = _norm_cyr(full_name).split()
    if not parts:
        return ""
    sur = parts[0].upper()
    ini = ""
    if len(parts) >= 2 and parts[1]:
        ini += parts[1][0].upper() + "."
    if len(parts) >= 3 and parts[2]:
        ini += parts[2][0].upper() + "."
    return f"{sur} {ini}".strip()

def doctor_key_initials(initials_name: str) -> str:
    """'Иванов И.П.' or 'Иванов И. П.' -> normalized key"""
    s = _norm_cyr(initials_name).upper()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("..",".")
    # ensure format SUR I.P.
    m = re.match(r"^([А-ЯA-Z\-]+)\s+([А-ЯA-Z])\.\s*([А-ЯA-Z])\.$", s)
    if m:
        return f"{m.group(1)} {m.group(2)}.{m.group(3)}."
    # fallback: keep as is
    return s

def _extract_dates(text: str, year: int, month: int) -> Set[dt.date]:
    """Extract DD.MM or DD.MM.YYYY or YYYY-MM-DD occurrences; keep only those matching month/year if provided."""
    out=set()
    if not text or not isinstance(text, str):
        return out
    t = _norm_cyr(text)
    # ranges like 01.10-12.10
    for m in re.finditer(r"(\d{1,2}\.\d{1,2})(?:\.\d{4})?\s*[-–—]\s*(\d{1,2}\.\d{1,2})(?:\.\d{4})?", t):
        a = parse_date(m.group(1), year, month)
        b = parse_date(m.group(2), year, month)
        for d in daterange(a,b):
            if d.year == year and d.month == month:
                out.add(d)
    # single dates
    for m in re.finditer(r"(\d{1,2}\.\d{1,2}(?:\.\d{4})?)", t):
        try:
            d = parse_date(m.group(1), year, month)
            if d.year == year and d.month == month:
                out.add(d)
        except Exception:
            pass
    for m in re.finditer(r"(\d{4}-\d{2}-\d{2})", t):
        try:
            d = parse_date(m.group(1), year, month)
            if d.year == year and d.month == month:
                out.add(d)
        except Exception:
            pass
    return out

def parse_wishlist_xlsx(file_bytes: bytes, doctors: List[Doctor], year: int, month: int) -> Dict:
    """
    Returns dict with:
      extra_ok_yes: set(full names)
      extra_ok_no: set(full names)
      add_vac: list[Vacation] (single-day or multi-day)
      date_off_hard: dict[name]->set[date]
      weekday_shift_pref: dict[name]->dict[weekday(0=Mon)]->'у'/'в'
      evenodd_shift_pref: dict[name]->('even'->'у'/'в', 'odd'->'у'/'в') optional
      priority_override: dict[name]->cabin
    """
    import pandas as _pd

    df = _pd.read_excel(BytesIO(file_bytes))
    # Build mapping from wishlist keys to full names
    full_by_key = {doctor_key_fullname(d.name): d.name for d in doctors}

    out = {
        "extra_ok_yes": set(),
        "extra_ok_no": set(),
        "add_vac": [],
        "date_off_hard": defaultdict(set),
        "weekday_shift_pref": defaultdict(dict),
        "evenodd_shift_pref": {},
        "priority_override": {},
    }

    # Column names in wishlist
    col_name = next((c for c in df.columns if "Фам" in str(c)), df.columns[0])
    col_sched = next((c for c in df.columns if "график" in str(c).lower()), None)
    col_vac = next((c for c in df.columns if "Даты" in str(c) and "отпуск" in str(c).lower()), None)
    col_extra = next((c for c in df.columns if "Дополнитель" in str(c)), None)
    col_prio = next((c for c in df.columns if "Приоритет" in str(c)), None)
    col_other = next((c for c in df.columns if "Другие" in str(c)), None)

    weekday_words = {
        "понедель":0, "вторник":1, "вторн":1, "сред":2, "четвер":3, "пятниц":4, "суббот":5, "воскрес":6
    }

    for _, row in df.iterrows():
        raw_name = str(row.get(col_name, "")).strip()
        if not raw_name or raw_name.lower() == "nan":
            continue
        key = doctor_key_initials(raw_name)
        full = full_by_key.get(key)
        if not full:
            continue

        # priority override cabin
        if col_prio is not None:
            cab = str(row.get(col_prio, "")).strip()
            cab = _norm_cyr(cab)
            if cab and cab.lower() != "nan":
                out["priority_override"][full] = cab

        # extra shifts consent
        if col_extra is not None:
            extra = str(row.get(col_extra, "")).strip()
            extra_n = _norm_cyr(extra).lower()
            if "не хочу" in extra_n or "не могу" in extra_n:
                out["extra_ok_no"].add(full)
            elif "могу" in extra_n or "подзаработ" in extra_n or "доп" in extra_n:
                out["extra_ok_yes"].add(full)

        # additional vacation dates
        if col_vac is not None:
            vac_text = row.get(col_vac, "")
            dates = _extract_dates(str(vac_text), year, month)
            for d in sorted(dates):
                out["add_vac"].append(Vacation(name=full, start=d, end=d))

        # schedule patterns even/odd
        if col_sched is not None:
            sched_text = _norm_cyr(str(row.get(col_sched, ""))).lower()
            if "четн" in sched_text and "нечетн" in sched_text:
                # determine mapping
                # "четные утром, нечетные вечером" etc
                even = "у" if "четн" in sched_text and "утр" in sched_text.split("нечетн")[0] else None
                # simpler robust:
                if "четн" in sched_text and "утр" in sched_text and "нечетн" in sched_text and "вечер" in sched_text:
                    # check which comes first after "четные"
                    m1 = re.search(r"четн\w*\s*(утр|вечер)", sched_text)
                    m2 = re.search(r"нечетн\w*\s*(утр|вечер)", sched_text)
                    if m1 and m2:
                        even = "у" if m1.group(1).startswith("утр") else "в"
                        odd = "у" if m2.group(1).startswith("утр") else "в"
                        out["evenodd_shift_pref"][full] = {"even": even, "odd": odd}

        # other wishes: hard days off, weekday shift restrictions (soft by reward later)
        if col_other is not None:
            other = _norm_cyr(str(row.get(col_other, ""))).lower()
            if other and other != "nan":
                dates = _extract_dates(other, year, month)
                # if text indicates off or "не ставить"
                if ("выходн" in other) or ("не став" in other) or ("не работать" in other):
                    for d in dates:
                        out["date_off_hard"][full].add(d)

                # weekday shift preferences in text
                for ww, wdi in weekday_words.items():
                    if ww in other:
                        # look near that word for "1" or "2" or "первая" "вторая"
                        # crude but useful
                        # find segment around weekday word
                        seg = other
                        # shift 1 -> morning, shift 2 -> evening
                        pref = None
                        if "1" in seg or "перв" in seg:
                            pref = "у"
                        if "2" in seg or "втор" in seg:
                            # avoid confusion with 'вторник'; require '2' or 'вторую смен'
                            if "2" in seg or "вторую смен" in seg:
                                pref = "в"
                        if pref in ("у","в"):
                            out["weekday_shift_pref"][full][wdi] = pref

    return out

# ----------------------------
# Priority rule: 3 doctors same cabin -> one loses priority
# ----------------------------

def apply_priority_collision_rule(doctors: List[Doctor]) -> Tuple[List[Doctor], List[Tuple[str,str]]]:
    """
    If >=3 doctors share the same TOP priority cabin, choose one deterministically and remove ALL priorities for them.
    Deterministic selection:
      1) Prefer removing a half-time doctor (0.5) if present (their priority already lower)
      2) Otherwise, lexicographically last name (stable).
    Returns updated doctors and list of (doctor, cabin) removed.
    """
    top_map = defaultdict(list)
    for d in doctors:
        if d.priorities:
            top_map[d.priorities[0]].append(d)

    removed = []
    for cabin, ds in top_map.items():
        if len(ds) >= 3:
            half = [d for d in ds if d.fte < 0.99]
            if half:
                pick = sorted(half, key=lambda x: x.name)[-1]
            else:
                pick = sorted(ds, key=lambda x: x.name)[-1]
            removed.append((pick.name, cabin))
            pick.priorities = []  # "для него не будет приоритетного кабинета"
    return doctors, removed


# ----------------------------
# Norm and calendar
# ----------------------------

def required_norm(doc: Doctor, workdays: List[dt.date], vac_days: Set[dt.date]) -> int:
    base = 22 if doc.fte >= 0.99 else 11
    vac_wd = sum(1 for d in workdays if d in vac_days)
    return max(0, base - vac_wd)


def build_slots(days: List[dt.date], cabins: List[str], holidays: Set[dt.date]):
    """
    Returns list of slots:
      (day_idx, date, shift_code, cabin)
    shift_code: 'у','в','р'
    """
    slots = []
    for di, day in enumerate(days):
        wkend = is_weekend(day) or (day in holidays)
        shifts = ['р'] if wkend else ['у','в']
        for sh in shifts:
            for cab in cabins:
                slots.append((di, day, sh, cab))
    return slots


# ----------------------------
# OR-Tools CP-SAT scheduler
# ----------------------------

def solve_with_cpsat(
    doctors: List[Doctor],
    vacations: List[Vacation],
    cabins: List[str],
    year: int,
    month: int,
    holidays: Set[dt.date],
    extra_ok: Set[str],
    wish: Optional[Dict] = None,
    shift_pref: Optional[Dict[str, Optional[str]]] = None,
    pref_weight: int = 3,
    enforce_morning_evening_mix: bool = True,
    extra_max: int = 6,
    time_limit_s: int = 15,
    free_label: str = "свободно",
):
    days = all_days_in_month(year, month)
    workdays = working_days_in_month(year, month)
    vac_map = build_vac_map(vacations)

    # Norms
    norm = {d.name: required_norm(d, workdays, vac_map.get(d.name,set())) for d in doctors}

    # Apply collision rule (>=3 share top cabin)
    doctors, removed = apply_priority_collision_rule(doctors)

    # Indices
    doc_names = [d.name for d in doctors]
    doc_idx = {n:i for i,n in enumerate(doc_names)}
    slots = build_slots(days, cabins, holidays)
    slot_idx = {(di, sh, cab): si for si,(di,_,sh,cab) in enumerate(slots)}

    model = cp_model.CpModel()

    # Decision vars: x[si, dj] = 1 if doctor dj assigned to slot si
    x = {}
    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, name in enumerate(doc_names):
            # vacation constraint: forbid
            if day in vac_map.get(name, set()):
                continue
            x[(si,dj)] = model.NewBoolVar(f"x_s{si}_d{dj}")

    # Free slot var
    free = [model.NewBoolVar(f"free_{si}") for si in range(len(slots))]

    # Each slot: exactly one (doctor or free)
    for si in range(len(slots)):
        vars_in = [x[(si,dj)] for dj in range(len(doc_names)) if (si,dj) in x]
        model.Add(sum(vars_in) + free[si] == 1)

    # One doctor per day max 1 shift: sum over all cabins and shifts that day <=1
    # Build work var per (day,doc)
    work = {}
    for di, day in enumerate(days):
        for dj in range(len(doc_names)):
            w = model.NewBoolVar(f"work_{di}_{dj}")
            work[(di,dj)] = w
            vars_in = []
            for cab in cabins:
                for sh in (['р'] if (is_weekend(day) or (day in holidays)) else ['у','в']):
                    si = slot_idx[(di, sh, cab)]
                    if (si,dj) in x:
                        vars_in.append(x[(si,dj)])
            if vars_in:
                model.Add(sum(vars_in) == w)
            else:
                model.Add(w == 0)

    for di, day in enumerate(days):
        for dj in range(len(doc_names)):
            # also ensures <=1 because w is 0/1 and equals sum of assignments
            pass

    # Hard wish-list days off (if provided): forbid working that day
    if wish and wish.get("date_off_hard"):
        for dj, name in enumerate(doc_names):
            for di, day in enumerate(days):
                if day in wish["date_off_hard"].get(name, set()):
                    model.Add(work[(di,dj)] == 0)

# Max 5 consecutive working days: in any window of 6 days, <=5 worked
    for dj in range(len(doc_names)):
        for start in range(0, len(days)-6+1):
            model.Add(sum(work[(di,dj)] for di in range(start, start+6)) <= 5)

# Enforce that doctors do not work only mornings or only evenings across the month (weekdays)
    # If a doctor has >=2 weekday shifts total, require at least one morning and one evening.
    if enforce_morning_evening_mix:
        for dj, name in enumerate(doc_names):
            weekday_m = []
            weekday_e = []
            for di, day in enumerate(days):
                if is_weekend(day) or (day in holidays):
                    continue
                for cab in cabins:
                    si_m = slot_idx.get((di, 'у', cab))
                    si_e = slot_idx.get((di, 'в', cab))
                    if si_m is not None and (si_m, dj) in x:
                        weekday_m.append(x[(si_m, dj)])
                    if si_e is not None and (si_e, dj) in x:
                        weekday_e.append(x[(si_e, dj)])

            if not weekday_m and not weekday_e:
                continue

            m_cnt = sum(weekday_m) if weekday_m else 0
            e_cnt = sum(weekday_e) if weekday_e else 0
            tot_wd = m_cnt + e_cnt

            has_two_or_more = model.NewBoolVar(f"has_two_weekday_{dj}")
            model.Add(tot_wd >= 2).OnlyEnforceIf(has_two_or_more)
            model.Add(tot_wd <= 1).OnlyEnforceIf(has_two_or_more.Not())
            model.Add(m_cnt >= 1).OnlyEnforceIf(has_two_or_more)
            model.Add(e_cnt >= 1).OnlyEnforceIf(has_two_or_more)

    # Monthly norm constraints:
    # For doctors without consent: exact norm (can be relaxed to <=norm if infeasible; we keep exact by default)
    # For extra_ok doctors: allow [norm, norm+extra_max]
    for dj, name in enumerate(doc_names):
        total_work = sum(work[(di,dj)] for di in range(len(days)))
        if name in extra_ok:
            model.Add(total_work >= norm[name])
            model.Add(total_work <= norm[name] + extra_max)
        else:
            model.Add(total_work == norm[name])

    # Objective: minimize free slots, and reward priority cabins.
    # Lower priority weight for half-time doctors.
    pri_weight_full = 8
    pri_weight_half = 3  # lower
    fill_weight = 20

    obj_terms = []
    # minimize free: maximize (1 - free)
    for si in range(len(slots)):
        obj_terms.append(fill_weight * (1 - free[si]))

    # reward priorities
    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, doc in enumerate(doctors):
            if (si,dj) not in x:
                continue
            if cab in doc.priorities:
                w = pri_weight_half if doc.fte < 0.99 else pri_weight_full
                obj_terms.append(w * x[(si,dj)])

    # wish-list shift preferences (soft): reward matching preferred shift on given days
    wish_weight = 2
    if wish and wish.get("pref_shift"):
        pref_shift = wish["pref_shift"]  # dict[(di,dj)] -> 'у'/'в'
        for (di, dj), pref in pref_shift.items():
            if pref not in ("у", "в"):
                continue
            day = days[di]
            if is_weekend(day) or (day in holidays):
                continue
            vars_pref = []
            for cab in cabins:
                si = slot_idx[(di, pref, cab)]
                if (si, dj) in x:
                    vars_pref.append(x[(si, dj)])
            if vars_pref:
                obj_terms.append(wish_weight * sum(vars_pref))

        # Doctor-level preferred shift (soft): reward assignments that match the doctor's preferred weekday shift.
    # shift_pref: dict name -> 'у'/'в'/None
    if shift_pref:
        for dj, doc in enumerate(doctors):
            pref = shift_pref.get(doc.name)
            if pref not in ('у', 'в'):
                continue
            for di, day in enumerate(days):
                if is_weekend(day) or (day in holidays):
                    continue
                vars_match = []
                for cab in cabins:
                    si = slot_idx[(di, pref, cab)]
                    if (si, dj) in x:
                        vars_match.append(x[(si, dj)])
                if vars_match:
                    obj_terms.append(int(pref_weight) * sum(vars_match))

    model.Maximize(sum(obj_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("CP-SAT не нашёл допустимое решение (попробуйте увеличить time_limit или разрешить extra_ok).")

    # Build schedule per doctor/day
    sched: Dict[str, Dict[dt.date, Tuple[str, str]]] = {n:{} for n in doc_names}
    for name in doc_names:
        for day in days:
            if day in vac_map.get(name,set()):
                sched[name][day] = ("от","")
            else:
                sched[name][day] = ("-","")

    slot_assign: Dict[dt.date, List[Tuple[str,str,str]]] = {d: [] for d in days}

    for si, (di, day, sh, cab) in enumerate(slots):
        who = free_label
        for dj, name in enumerate(doc_names):
            if (si,dj) in x and solver.Value(x[(si,dj)]) == 1:
                who = name
                break
        slot_assign[day].append((sh,cab,who))
        if who != free_label:
            sched[who][day] = (sh,cab)

    deviation = {}
    for name in doc_names:
        fact = sum(1 for day in days if sched[name][day][0] in ("у","в","р"))
        deviation[name] = fact - norm[name]

    meta = {
        "removed_priorities": removed,
        "objective": solver.ObjectiveValue(),
        "status": "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE",
    }
    return days, norm, sched, slot_assign, deviation, meta


# ----------------------------
# XLSX export to bytes (6 sheets)
# ----------------------------

def export_xlsx_bytes(doctors: List[Doctor],
                      days: List[dt.date],
                      sched: Dict[str, Dict[dt.date, Tuple[str, str]]],
                      norm: Dict[str, int],
                      slot_assign: Dict[dt.date, List[Tuple[str, str, str]]],
                      cabins: List[str],
                      free_label: str = "свободно") -> bytes:

    wb = Workbook()
    ws = wb.active
    ws.title = "График врачей"

    fill_morning = PatternFill("solid", fgColor="A7D8FF")  # голубой
    fill_evening = PatternFill("solid", fgColor="FFB6C1")  # розовый
    fill_weekend = PatternFill("solid", fgColor="A7F3A7")  # зелёный
    fill_vac = PatternFill("solid", fgColor="C0C0C0")      # серый
    fill_off = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Врач").font = header_font
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=d.day)
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 11

    last_col = 2 + len(days)
    ws.cell(row=1, column=last_col, value="Смен").font = header_font
    ws.cell(row=1, column=last_col + 1, value="Комментарий").font = header_font

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(last_col)].width = 8
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 28

    for r, doc in enumerate(doctors, start=2):
        ws.cell(row=r, column=1, value=doc.name).border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)

        shift_count = 0
        for i, day in enumerate(days, start=2):
            code, cab = sched[doc.name][day]
            txt = code
            if code in ("у", "в", "р"):
                txt = f"{code} ({cab})"
                shift_count += 1

            cell = ws.cell(row=r, column=i, value=txt)
            cell.alignment = center
            cell.border = border

            if code == "у":
                cell.fill = fill_morning
            elif code == "в":
                cell.fill = fill_evening
            elif code == "р":
                cell.fill = fill_weekend
            elif code == "от":
                cell.fill = fill_vac
            else:
                cell.fill = fill_off

        ws.cell(row=r, column=last_col, value=shift_count).alignment = center
        ws.cell(row=r, column=last_col).border = border

        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        ws.cell(row=r, column=last_col + 1, value=f"ставка {rate}, норма {norm[doc.name]}").border = border
        ws.cell(row=r, column=last_col + 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("Сводка по врачам")
    ws2.append(["Врач", "Ставка", "Норма смен", "Факт смен", "Отклонение", "Приоритеты"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        ws2.append([doc.name, doc.fte, norm[doc.name], fact, fact - norm[doc.name], ", ".join(doc.priorities)])

    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=6):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["F"].width = 26

    # Sheet 3
    ws3 = wb.create_sheet("Загрузка кабинетов")
    ws3.append(["Дата", "День", "Смена", "Кабинет", "Врач"])
    for c in ws3[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for d in days:
        for sh, cab, who in slot_assign[d]:
            ws3.append([d.isoformat(), day_name[d.weekday()], sh, cab, who])

    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 6
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 34

    # Sheet 4
    ws4 = wb.create_sheet("Общая статистика")
    ws4.append(["Показатель", "Значение"])
    ws4["A1"].font = header_font
    ws4["B1"].font = header_font

    total_slots = sum(len(slot_assign[d]) for d in days)
    free_slots = sum(1 for d in days for sh, cab, who in slot_assign[d] if who == free_label)
    ws4.append(["Месяц", f"{days[0].strftime('%B')} {days[0].year}"])
    ws4.append(["Кабинетов", len(cabins)])
    ws4.append(["Всего слотов (кабинет-смена)", total_slots])
    ws4.append(["Свободно (не заполнено)", free_slots])
    ws4.append(["Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0])

    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18

    # Sheet 5
    ws5 = wb.create_sheet("Обоснование отклонений")
    ws5.append(["Врач", "Норма", "Факт", "Отклонение", "Пояснение"])
    for c in ws5[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        dev = fact - norm[doc.name]
        vac_count = sum(1 for d in days if sched[doc.name][d][0] == "от")
        expl = []
        if vac_count:
            expl.append(f"отпуск: {vac_count} дн.")
        if dev > 0:
            expl.append(f"доп. смены: {dev}")
        elif dev < 0:
            expl.append(f"не добрали: {-dev}")
        else:
            expl.append("норма выполнена")
        ws5.append([doc.name, norm[doc.name], fact, dev, "; ".join(expl)])

    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["E"].width = 42

    # Sheet 6
    ws6 = wb.create_sheet("Кабинетов на врача")
    ws6.append(["Врач", "Кабинет", "Кол-во смен"])
    for c in ws6[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        cab_counter = Counter()
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                cab_counter[cab] += 1
        for cab, cnt in cab_counter.most_common():
            ws6.append([doc.name, cab, cnt])

    for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws6.column_dimensions["A"].width = 34
    ws6.column_dimensions["B"].width = 10
    ws6.column_dimensions["C"].width = 14

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------

st.set_page_config(page_title="График смен врачей (CP-SAT)", layout="wide")
st.title("График смен врачей — OR-Tools CP-SAT")

with st.sidebar:
    st.header("Параметры")
    year = st.number_input("Год", min_value=2020, max_value=2035, value=2025, step=1)
    month = st.number_input("Месяц", min_value=1, max_value=12, value=10, step=1)
    time_limit = st.slider("Time limit (сек)", min_value=3, max_value=60, value=15, step=1)
    extra_max = st.slider("Максимум доп.смен для согласившихся", min_value=0, max_value=12, value=6, step=1)

    use_wish = st.checkbox("Использовать wish list", value=True)
    wish_file = st.file_uploader("wish_list.xlsx", type=["xlsx"], disabled=not use_wish)

    enforce_mix = st.checkbox("Требовать смесь утро/вечер (если ≥2 будних смен)", value=True)
    pref_weight = st.slider("Вес предпочтения смены (утро/вечер) в оптимизации", min_value=0, max_value=10, value=3, step=1)

    st.caption("CSV форматы; комментарии начинаются с #. Даты: DD.MM.YYYY или YYYY-MM-DD.")

    time_limit_s = st.number_input(
        "Лимит времени оптимизации (сек)",
        min_value=5,
        max_value=300,
        value=30,
        step=5
    )

    default_doctors = """\
Боброва Эльвира Анатольевна,1
Буданова Татьяна Владимировна,0.5
Варвус Иван Михайлович,1
Гагкаева Мария Аркадьевна,1
Гущина Юлия Викторовна,1
Джафаров Башир Темирханович,0.5
Догузова Александра Гочевна,1
Евдокимова Жанна Владимировна,1
Егуренкова Виктория Сергеевна,1
Захаров Александр Сергеевич,1
Исаев Абсалам Темурланович,1
Китков Игорь Игоревич,1
Кокорина Серафима Евгеньевна,1
Королькова Александра Викторовна,0.5
Короткова Полина Александровна,1
Лукьяненко Владимир Александрович,1
Маркин Александр Андреевич,1
Марфутов Василий Васильевич,1
Медведева Мария Вячеславовна,1
Найданова Лидия Вадимовна,1
Нам Ирина Николаевна,1
Петрова Полина Юрьевна,1
Пономарева Юлия Константиновна,1
Прокофьев Евгений Андреевич,1
Пусь Юлия Владимировна,1
Танкиева Фатима Умат-Гиреевна,1
Тетерев Ярослав Тарасович,1
Федорова Оксана Сергеевна,1
Харитонова Наталья Игоревна,1
Шепель Артем Олегович,1
Юдин Алексей Александрович,1
Юдина Виктория Дмитриевна,1
"""
    default_priorities = """\
Пономарева Юлия Константиновна,2А03
Лукьяненко Владимир Александрович,2А03
Королькова Александра Викторовна,2А03
Харитонова Наталья Игоревна,2А04
Исаев Абсалам Темурланович,2А04
Короткова Полина Александровна,2А04
Гущина Юлия Викторовна,2А05
Федорова Оксана Сергеевна,2А05
Кокорина Серафима Евгеньевна,2А05
Егуренкова Виктория Сергеевна,2А06
Евдокимова Жанна Владимировна,2А06
Буданова Татьяна Владимировна,2А06
Боброва Эльвира Анатольевна,2А07
Марфутов Василий Васильевич,2А07
Догузова Александра Гочевна,2А07
Медведева Мария Вячеславовна,2А08
Юдин Алексей Александрович,2А08
Юдина Виктория Дмитриевна,2А33
Шепель Артем Олегович,2А33
Китков Игорь Игоревич,2А34
Гагкаева Мария Аркадьевна,2А34
Варвус Иван Михайлович,2А49
Маркин Александр Андреевич,2А49
Тетерев Ярослав Тарасович,2А52
Джафаров Башир Темирханович,2А52
Петрова Полина Юрьевна,2А52
Пусь Юлия Владимировна,2А54
Прокофьев Евгений Андреевич,2А54
Захаров Александр Сергеевич,2А54
Найданова Лидия Вадимовна,2А55
Нам Ирина Николаевна,2А55|2А54
Танкиева Фатима Умат-Гиреевна,2А55
"""
    default_vac = """\
Боброва Эльвира Анатольевна,01.10.2025,12.10.2025
Тетерев Ярослав Тарасович,29.09.2025,12.10.2025
Исаев Абсалам Темурланович,06.10.2025,19.10.2025
Гагкаева Мария Аркадьевна,01.10.2025,14.10.2025
Гагкаева Мария Аркадьевна,20.10.2025,20.10.2025
Гагкаева Мария Аркадьевна,27.10.2025,27.10.2025
Егуренкова Виктория Сергеевна,06.10.2025,19.10.2025
"""
    default_cabins = "2А03, 2А04, 2А05, 2А06, 2А07, 2А08, 2А33, 2А34, 2А49, 2А52, 2А54, 2А55"
    default_holidays = ""
    default_extra_ok = ""  # names who agree to take extra shifts
    default_shift_pref = ""  # CSV: "ФИО,у" or "ФИО,в" or "ФИО,нет"

    doctors_text = st.text_area("Врачи (name,fte)", value=default_doctors, height=240)
    prio_text = st.text_area("Приоритеты (name,cab1|cab2)", value=default_priorities, height=240)
    vac_text = st.text_area("Отпуска (name,start,end)", value=default_vac, height=160)
    cabins_text = st.text_area("Кабинеты", value=default_cabins, height=70)
    holidays_text = st.text_area("Праздники/доп. выходные (опц.)", value=default_holidays, height=60)
    extra_ok_text = st.text_area("Согласны на доп. смены (по одному имени в строке, опц.)", value=default_extra_ok, height=90)
    shift_pref_text = st.text_area("Предпочтение смены (CSV: ФИО,у/в/нет) — мягкое", value=default_shift_pref, height=110)

    recompute_clicked = st.button("Recompute (CP-SAT)", type="primary", use_container_width=True)

# Persist computed state
if "computed" not in st.session_state:
    st.session_state.computed = False

def compute(use_wish, wish_file):
    wish = None
    if use_wish and wish_file is not None:
        wish = wish_file.getvalue()

    doctors = parse_doctors_csv(doctors_text)
    pr_map = parse_priorities_csv(prio_text)
    for d in doctors:
        d.priorities = pr_map.get(d.name, [])

    vacations = parse_vacations_csv(vac_text, int(year), int(month))
    cabins = parse_cabins(cabins_text)
    holidays = parse_holidays(holidays_text, int(year), int(month))
    extra_ok = parse_yes_list(extra_ok_text)
    shift_pref = parse_shift_pref_csv(shift_pref_text)

    if not doctors:
        raise ValueError("Список врачей пуст.")
    if not cabins:
        raise ValueError("Список кабинетов пуст.")

    wish = None
    if use_wish:
        wish_bytes = None
        if wish_file is not None:
            wish_bytes = wish_file.getvalue()
        else:
            # If file exists рядом с приложением, попробуем прочитать
            try:
                with open("wish_list.xlsx", "rb") as f:
                    wish_bytes = f.read()
            except Exception:
                wish_bytes = None

        if wish_bytes:
            wish = parse_wishlist_xlsx(wish_bytes, doctors, int(year), int(month))

            # Apply priority override from wish list
            for d in doctors:
                cab = wish.get("priority_override", {}).get(d.name)
                if cab:
                    if cab in d.priorities:
                        d.priorities.remove(cab)
                    d.priorities.insert(0, cab)

            # Add extra vacations from wish list
            add_v = wish.get("add_vac", [])
            if add_v:
                vacations.extend(add_v)

            # Update extra_ok from wish list consent flags
            extra_ok |= set(wish.get("extra_ok_yes", set()))
            extra_ok -= set(wish.get("extra_ok_no", set()))

            # Build preferred shift map for objective: (di,dj) -> 'у' or 'в'
            days_for_map = all_days_in_month(int(year), int(month))
            doc_names = [d.name for d in doctors]
            name_to_dj = {n:i for i,n in enumerate(doc_names)}
            pref_shift = {}

            evenodd = wish.get("evenodd_shift_pref", {})
            weekday_pref = wish.get("weekday_shift_pref", {})
            for di, day in enumerate(days_for_map):
                if is_weekend(day) or (day in holidays):
                    continue
                for name in doc_names:
                    dj = name_to_dj[name]
                    pref = None
                    # weekday preference has priority
                    wd = day.weekday()
                    if name in weekday_pref and wd in weekday_pref[name]:
                        pref = weekday_pref[name][wd]
                    elif name in evenodd:
                        pref = evenodd[name]["even"] if (day.day % 2 == 0) else evenodd[name]["odd"]
                    if pref in ("у","в"):
                        pref_shift[(di,dj)] = pref
            wish["pref_shift"] = pref_shift

    days, norm, sched, slot_assign, deviation, meta = solve_with_cpsat(
        doctors=doctors,
        vacations=vacations,
        cabins=cabins,
        year=int(year),
        month=int(month),
        holidays=holidays,
        extra_ok=extra_ok,
        wish=wish,
        extra_max=int(extra_max),
        time_limit_s=int(time_limit_s),
        shift_pref=shift_pref,
        pref_weight=int(pref_weight),
        enforce_morning_evening_mix=bool(enforce_mix),
    )

    st.session_state.days = days
    st.session_state.doctors = doctors
    st.session_state.vacations = vacations
    st.session_state.cabins = cabins
    st.session_state.holidays = holidays
    st.session_state.extra_ok = extra_ok
    st.session_state.wish = wish
    st.session_state.norm = norm
    st.session_state.sched = sched
    st.session_state.slot_assign = slot_assign
    st.session_state.deviation = deviation
    st.session_state.meta = meta
    st.session_state.computed = True


if recompute_clicked or not st.session_state.computed:
    try:
        with st.spinner("CP-SAT решает задачу..."):
            compute(use_wish, wish_file)
        st.success("Готово.")
    except Exception as e:
        st.session_state.computed = False
        st.error(f"Ошибка: {e}")

if st.session_state.computed:
    days: List[dt.date] = st.session_state.days
    doctors: List[Doctor] = st.session_state.doctors
    norm = st.session_state.norm
    sched = st.session_state.sched
    slot_assign = st.session_state.slot_assign
    cabins = st.session_state.cabins
    meta = st.session_state.meta

    # Info & compliance notes
    colA, colB = st.columns([2,3])
    with colA:
        st.subheader("Смены и времена")
        st.markdown(
            "- **Будни:** у (7:30–13:30), в (14:20–20:20)\n"
            "- **Выходные/праздники:** р (9:00–18:00)\n"
            "- **Ограничения:** ≤1 смена/день, ≤5 дней подряд, отпуск = от"
        )
    with colB:
        st.subheader("Решение CP-SAT")
        st.write({"status": meta["status"], "objective": meta["objective"]})
        if meta["removed_priorities"]:
            st.warning("Правило '3 врача на 1 кабинет': одному врачу снят приоритет.")
            st.dataframe(pd.DataFrame(meta["removed_priorities"], columns=["Врач", "Кабинет (снят приоритет)"]),
                         use_container_width=True, height=180)

    # Main schedule table (doctor x day)
    rows = []
    for doc in doctors:
        row = {"Врач": doc.name}
        shift_count = 0
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                row[str(d.day)] = f"{code} ({cab})"
                shift_count += 1
            else:
                row[str(d.day)] = code
        row["Смен"] = shift_count
        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        row["Комментарий"] = f"ставка {rate}, норма {norm[doc.name]}"
        rows.append(row)

    df = pd.DataFrame(rows)
    st.subheader("Лист 1 — График врачей")
    st.dataframe(df, use_container_width=True, height=520)

    # Summary
    st.subheader("Сводка по врачам")
    summ_rows = []
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у","в","р"))
        summ_rows.append({
            "Врач": doc.name,
            "Ставка": doc.fte,
            "Норма смен": norm[doc.name],
            "Факт смен": fact,
            "Отклонение": fact - norm[doc.name],
            "Приоритеты": ", ".join(doc.priorities),
        })
    st.dataframe(pd.DataFrame(summ_rows), use_container_width=True, height=320)

    # Download XLSX
    xlsx_bytes = export_xlsx_bytes(
        doctors=doctors,
        days=days,
        sched=sched,
        norm=norm,
        slot_assign=slot_assign,
        cabins=cabins,
    )
    filename = f"график_{int(year)}_{int(month):02d}_cpsat.xlsx"
    st.download_button(
        label="Скачать XLSX (6 листов, цветной)",
        data=xlsx_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    with st.expander("Загрузка кабинетов (детально)"):
        load_rows = []
        day_name = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
        for d in days:
            for sh, cab, who in slot_assign[d]:
                load_rows.append({
                    "Дата": d.isoformat(),
                    "День": day_name[d.weekday()],
                    "Смена": sh,
                    "Кабинет": cab,
                    "Врач": who,
                })
        st.dataframe(pd.DataFrame(load_rows), use_container_width=True, height=420)
