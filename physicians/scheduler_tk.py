#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scheduler Tkinter UI (Oct 2025 default) + constraint satisfaction style assignment.

Fixes Tkinter error:
    _tkinter.TclError: can't pack ... inside ...
by ensuring every widget is created with its final parent (no pack(in_=...) reparenting).

Deps:
    pip install openpyxl

Run:
    python scheduler_tk.py
"""

import re
import random
import datetime as dt
from dataclasses import dataclass, field
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Set, Optional

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# Data model
# ----------------------------

@dataclass
class Doctor:
    name: str
    fte: float = 1.0  # 1.0 or 0.5
    priorities: List[str] = field(default_factory=list)

@dataclass
class Vacation:
    name: str
    start: dt.date
    end: dt.date  # inclusive


# ----------------------------
# Parsing helpers
# ----------------------------

def parse_date(s: str, year: int, month: int) -> dt.date:
    s = s.strip()
    # DD.MM or DD.MM.YYYY or YYYY-MM-DD
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        d, m, y = map(int, s.split("."))
        return dt.date(y, m, d)
    if re.match(r"^\d{2}\.\d{2}$", s):
        d, m = map(int, s.split("."))
        return dt.date(year, m, d)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        y, m, d = map(int, s.split("-"))
        return dt.date(y, m, d)
    raise ValueError(f"Bad date: {s}")

def daterange(a: dt.date, b: dt.date):
    cur = a
    while cur <= b:
        yield cur
        cur += dt.timedelta(days=1)

def is_weekend(d: dt.date) -> bool:
    return d.weekday() >= 5

def all_days_in_month(year: int, month: int) -> List[dt.date]:
    days = []
    d = dt.date(year, month, 1)
    while d.month == month:
        days.append(d)
        d += dt.timedelta(days=1)
    return days

def working_days_in_month(year: int, month: int) -> List[dt.date]:
    return [d for d in all_days_in_month(year, month) if not is_weekend(d)]

def build_vac_map(vacs: List[Vacation]) -> Dict[str, Set[dt.date]]:
    m = defaultdict(set)
    for v in vacs:
        for day in daterange(v.start, v.end):
            m[v.name].add(day)
    return m

def parse_doctors_csv(text: str) -> List[Doctor]:
    """
    CSV lines: name, fte(1 or 0.5)
    """
    out = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 1:
            continue
        name = parts[0]
        fte = float(parts[1]) if len(parts) >= 2 and parts[1] else 1.0
        out.append(Doctor(name=name, fte=fte))
    return out

def parse_priorities_csv(text: str) -> Dict[str, List[str]]:
    """
    CSV lines: name, cabin1|cabin2|...
    """
    out: Dict[str, List[str]] = {}
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        if len(parts) < 2:
            continue
        name = parts[0]
        cabs = [c.strip() for c in parts[1].split("|") if c.strip()]
        out[name] = cabs
    return out

def parse_vacations_csv(text: str, year: int, month: int) -> List[Vacation]:
    """
    CSV lines: name, start, end   (dates can be DD.MM or DD.MM.YYYY or YYYY-MM-DD)
    """
    out = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        name = parts[0]
        start = parse_date(parts[1], year, month)
        end = parse_date(parts[2], year, month)
        out.append(Vacation(name=name, start=start, end=end))
    return out

def parse_cabins(text: str) -> List[str]:
    items = re.split(r"[,\s]+", text.strip())
    return [i for i in (x.strip() for x in items) if i]

def parse_holidays(text: str, year: int, month: int) -> Set[dt.date]:
    out=set()
    if not text.strip():
        return out
    for tok in re.split(r"[,\s]+", text.strip()):
        if not tok.strip():
            continue
        out.add(parse_date(tok.strip(), year, month))
    return out


# ----------------------------
# CSP-like scheduler (heuristic with hard constraints)
# ----------------------------

def calc_required_shifts(doc: Doctor, workdays: List[dt.date], vac_days: Set[dt.date]) -> int:
    base = 23 if doc.fte >= 0.99 else 11
    vac_wd = sum(1 for d in workdays if d in vac_days)
    return max(0, base - vac_wd)

def score_candidate(doc: Doctor, cabin: str, shift_code: str,
                    remaining_need: int, streak: int,
                    used_today: bool, on_vac: bool,
                    weekend_total: int) -> float:
    # hard constraints
    if used_today or on_vac:
        return -1e9
    if streak >= 6:
        return -1e9

    s = 0.0
    s += remaining_need * 10.0                 # satisfy norm
    s -= max(0, -remaining_need) * 20.0        # avoid overtime
    if cabin in doc.priorities:
        s += 8.0
    if doc.fte < 1.0:
        s -= 3.0                               # half-time lower priority
    if shift_code == "р":
        s -= weekend_total * 1.0               # weekend fairness
    return s

def schedule_month(doctors: List[Doctor],
                   vacations: List[Vacation],
                   cabins: List[str],
                   year: int,
                   month: int,
                   holidays: Optional[Set[dt.date]] = None,
                   seed: int = 42,
                   free_label: str = "свободно"):

    random.seed(seed)
    days = all_days_in_month(year, month)
    workdays = working_days_in_month(year, month)
    holidays = holidays or set()
    vac_map = build_vac_map(vacations)

    required = {doc.name: calc_required_shifts(doc, workdays, vac_map.get(doc.name, set())) for doc in doctors}

    total = Counter()
    weekend_total = Counter()
    streak = Counter()

    sched: Dict[str, Dict[dt.date, Tuple[str, str]]] = {d.name: {} for d in doctors}
    for doc in doctors:
        for day in days:
            if day in vac_map.get(doc.name, set()):
                sched[doc.name][day] = ("от", "")

    slot_assign: Dict[dt.date, List[Tuple[str, str, str]]] = {}

    cab_pop = Counter()
    for doc in doctors:
        for c in doc.priorities:
            cab_pop[c] += 1

    for day in days:
        wkend = is_weekend(day) or (day in holidays)
        shifts = ["р"] if wkend else ["у", "в"]
        slots = [(sh, cab) for sh in shifts for cab in cabins]
        slots.sort(key=lambda x: (-cab_pop[x[1]], x[0], x[1]))

        used_today: Set[str] = set()
        day_assign: List[Tuple[str, str, str]] = []

        for sh, cab in slots:
            best_name = None
            best_score = -1e18
            order = doctors[:]
            random.shuffle(order)

            for doc in order:
                name = doc.name
                on_vac = day in vac_map.get(name, set())
                rem = required[name] - total[name]
                sc = score_candidate(doc, cab, sh, rem, streak[name], (name in used_today), on_vac, weekend_total[name])
                if sc > best_score:
                    best_score = sc
                    best_name = name

            if best_name is None or best_score < -1e8:
                day_assign.append((sh, cab, free_label))
            else:
                day_assign.append((sh, cab, best_name))
                used_today.add(best_name)

        slot_assign[day] = day_assign

        worked_today = {}
        for sh, cab, who in day_assign:
            if who != free_label:
                worked_today[who] = (sh, cab)

        for doc in doctors:
            name = doc.name
            if day in vac_map.get(name, set()):
                streak[name] = 0
                continue
            if name in worked_today:
                sh, cab = worked_today[name]
                sched[name][day] = (sh, cab)
                total[name] += 1
                if sh == "р":
                    weekend_total[name] += 1
                streak[name] += 1
            else:
                sched[name][day] = ("-", "")
                streak[name] = 0

    deviation = {name: sum(1 for d in days if sched[name][d][0] in ("у", "в", "р")) - required[name] for name in required}
    return days, required, sched, slot_assign, deviation


# ----------------------------
# Excel export
# ----------------------------

def export_xlsx(path: str,
                doctors: List[Doctor],
                days: List[dt.date],
                sched: Dict[str, Dict[dt.date, Tuple[str, str]]],
                required: Dict[str, int],
                slot_assign: Dict[dt.date, List[Tuple[str, str, str]]],
                cabins: List[str],
                free_label: str = "свободно"):

    wb = Workbook()
    ws = wb.active
    ws.title = "График врачей"

    fill_morning = PatternFill("solid", fgColor="A7D8FF")  # голубой
    fill_evening = PatternFill("solid", fgColor="FFB6C1")  # розовый
    fill_weekend = PatternFill("solid", fgColor="A7F3A7")  # зелёный
    fill_vac = PatternFill("solid", fgColor="C0C0C0")      # серый
    fill_off = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Врач").font = header_font
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=d.day)
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 11

    last_col = 2 + len(days)
    ws.cell(row=1, column=last_col, value="Смен").font = header_font
    ws.cell(row=1, column=last_col + 1, value="Комментарий").font = header_font

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(last_col)].width = 8
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 24

    for r, doc in enumerate(doctors, start=2):
        ws.cell(row=r, column=1, value=doc.name).border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)

        shift_count = 0
        for i, day in enumerate(days, start=2):
            code, cab = sched[doc.name][day]
            txt = code
            if code in ("у", "в", "р"):
                txt = f"{code} ({cab})"
                shift_count += 1

            cell = ws.cell(row=r, column=i, value=txt)
            cell.alignment = center
            cell.border = border

            if code == "у":
                cell.fill = fill_morning
            elif code == "в":
                cell.fill = fill_evening
            elif code == "р":
                cell.fill = fill_weekend
            elif code == "от":
                cell.fill = fill_vac
            else:
                cell.fill = fill_off

        ws.cell(row=r, column=last_col, value=shift_count).alignment = center
        ws.cell(row=r, column=last_col).border = border

        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        ws.cell(row=r, column=last_col + 1, value=f"ставка {rate}, норма {required[doc.name]}").border = border
        ws.cell(row=r, column=last_col + 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("Сводка по врачам")
    ws2.append(["Врач", "Ставка", "Норма смен", "Факт смен", "Отклонение", "Приоритеты"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        ws2.append([doc.name, doc.fte, required[doc.name], fact, fact - required[doc.name], ", ".join(doc.priorities)])

    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=6):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["F"].width = 26

    # Sheet 3
    ws3 = wb.create_sheet("Загрузка кабинетов")
    ws3.append(["Дата", "День", "Смена", "Кабинет", "Врач"])
    for c in ws3[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for d in days:
        for sh, cab, who in slot_assign[d]:
            ws3.append([d.isoformat(), day_name[d.weekday()], sh, cab, who])

    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 6
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 34

    # Sheet 4
    ws4 = wb.create_sheet("Общая статистика")
    ws4.append(["Показатель", "Значение"])
    ws4["A1"].font = header_font
    ws4["B1"].font = header_font

    total_slots = sum(len(slot_assign[d]) for d in days)
    free_slots = sum(1 for d in days for sh, cab, who in slot_assign[d] if who == free_label)
    ws4.append(["Месяц", f"{days[0].strftime('%B')} {days[0].year}"])
    ws4.append(["Кабинетов", len(cabins)])
    ws4.append(["Всего слотов (кабинет-смена)", total_slots])
    ws4.append(["Свободно (не заполнено)", free_slots])
    ws4.append(["Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0])

    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18

    # Sheet 5
    ws5 = wb.create_sheet("Обоснование отклонений")
    ws5.append(["Врач", "Норма", "Факт", "Отклонение", "Пояснение"])
    for c in ws5[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        dev = fact - required[doc.name]
        vac_count = sum(1 for d in days if sched[doc.name][d][0] == "от")
        expl = []
        if vac_count:
            expl.append(f"отпуск: {vac_count} дн.")
        if dev > 0:
            expl.append(f"доп. смены: {dev}")
        elif dev < 0:
            expl.append(f"не добрали: {-dev}")
        else:
            expl.append("норма выполнена")
        ws5.append([doc.name, required[doc.name], fact, dev, "; ".join(expl)])

    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["E"].width = 38

    # Sheet 6
    ws6 = wb.create_sheet("Кабинетов на врача")
    ws6.append(["Врач", "Кабинет", "Кол-во смен"])
    for c in ws6[1]:
        c.font = header_font
        c.border = border
        c.alignment = center

    for doc in doctors:
        cab_counter = Counter()
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                cab_counter[cab] += 1
        for cab, cnt in cab_counter.most_common():
            ws6.append([doc.name, cab, cnt])

    for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws6.column_dimensions["A"].width = 34
    ws6.column_dimensions["B"].width = 10
    ws6.column_dimensions["C"].width = 14

    wb.save(path)


# ----------------------------
# UI
# ----------------------------

class SchedulerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("График смен врачей — CSP (Tkinter)")
        self.geometry("1400x780")

        self.year_var = tk.IntVar(value=2025)
        self.month_var = tk.IntVar(value=10)

        self.days: List[dt.date] = []
        self.doctors: List[Doctor] = []
        self.vacations: List[Vacation] = []
        self.cabins: List[str] = []
        self.holidays: Set[dt.date] = set()

        self.required: Dict[str, int] = {}
        self.sched: Dict[str, Dict[dt.date, Tuple[str, str]]] = {}
        self.slot_assign: Dict[dt.date, List[Tuple[str, str, str]]] = {}

        self._build_ui()
        self._set_initial_values()
        self.recompute()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Год:").pack(side="left")
        ttk.Spinbox(top, from_=2020, to=2035, textvariable=self.year_var, width=6).pack(side="left", padx=6)

        ttk.Label(top, text="Месяц:").pack(side="left")
        ttk.Spinbox(top, from_=1, to=12, textvariable=self.month_var, width=4).pack(side="left", padx=6)

        ttk.Button(top, text="Recompute", command=self.recompute).pack(side="left", padx=10)
        ttk.Button(top, text="Export XLSX", command=self.export).pack(side="left", padx=6)

        self.status_var = tk.StringVar(value="Готово")
        ttk.Label(top, textvariable=self.status_var).pack(side="right")

        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        left = ttk.Frame(paned, width=430)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=3)

        # ---- Notebook with input tabs
        nb = ttk.Notebook(left)
        nb.pack(fill="both", expand=True)

        tab_doctors = ttk.Frame(nb)
        tab_prio = ttk.Frame(nb)
        tab_vac = ttk.Frame(nb)
        tab_misc = ttk.Frame(nb)

        nb.add(tab_doctors, text="Врачи (CSV)")
        nb.add(tab_prio, text="Приоритеты (CSV)")
        nb.add(tab_vac, text="Отпуска (CSV)")
        nb.add(tab_misc, text="Кабинеты/Праздники")

        self.txt_doctors = tk.Text(tab_doctors, wrap="none")
        self.txt_priorities = tk.Text(tab_prio, wrap="none")
        self.txt_vacations = tk.Text(tab_vac, wrap="none")

        self.txt_doctors.pack(fill="both", expand=True)
        self.txt_priorities.pack(fill="both", expand=True)
        self.txt_vacations.pack(fill="both", expand=True)

        ttk.Label(tab_misc, text="Кабинеты (через запятую или пробел):").pack(anchor="w", padx=6, pady=(8, 0))
        self.txt_cabins = tk.Text(tab_misc, height=4, wrap="none")
        self.txt_cabins.pack(fill="x", padx=6, pady=6)

        ttk.Label(tab_misc, text="Праздники/доп. выходные (даты через запятую, опц.):").pack(anchor="w", padx=6, pady=(8, 0))
        self.txt_holidays = tk.Text(tab_misc, height=3, wrap="none")
        self.txt_holidays.pack(fill="x", padx=6, pady=6)

        # ---- Schedule table
        table_frame = ttk.Frame(right)
        table_frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.pack(side="left", fill="both", expand=True)

        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")

    def _set_initial_values(self):
        doctors_csv = """\
Боброва Эльвира Анатольевна,1
Буданова Татьяна Владимировна,0.5
Варвус Иван Михайлович,1
Гагкаева Мария Аркадьевна,1
Гущина Юлия Викторовна,1
Джафаров Башир Темирханович,0.5
Догузова Александра Гочевна,1
Евдокимова Жанна Владимировна,1
Егуренкова Виктория Сергеевна,1
Захаров Александр Сергеевич,1
Исаев Абсалам Темурланович,1
Китков Игорь Игоревич,1
Кокорина Серафима Евгеньевна,1
Королькова Александра Викторовна,0.5
Короткова Полина Александровна,1
Лукьяненко Владимир Александрович,1
Маркин Александр Андреевич,1
Марфутов Василий Васильевич,1
Медведева Мария Вячеславовна,1
Найданова Лидия Вадимовна,1
Нам Ирина Николаевна,1
Петрова Полина Юрьевна,1
Пономарева Юлия Константиновна,1
Прокофьев Евгений Андреевич,1
Пусь Юлия Владимировна,1
Танкиева Фатима Умат-Гиреевна,1
Тетерев Ярослав Тарасович,1
Федорова Оксана Сергеевна,1
Харитонова Наталья Игоревна,1
Шепель Артем Олегович,1
Юдин Алексей Александрович,1
Юдина Виктория Дмитриевна,1
"""
        priorities_csv = """\
Пономарева Юлия Константиновна,2А03
Лукьяненко Владимир Александрович,2А03
Королькова Александра Викторовна,2А03
Харитонова Наталья Игоревна,2А04
Исаев Абсалам Темурланович,2А04
Короткова Полина Александровна,2А04
Гущина Юлия Викторовна,2А05
Федорова Оксана Сергеевна,2А05
Кокорина Серафима Евгеньевна,2А05
Егуренкова Виктория Сергеевна,2А06
Евдокимова Жанна Владимировна,2А06
Буданова Татьяна Владимировна,2А06
Боброва Эльвира Анатольевна,2А07
Марфутов Василий Васильевич,2А07
Догузова Александра Гочевна,2А07
Медведева Мария Вячеславовна,2А08
Юдин Алексей Александрович,2А08
Юдина Виктория Дмитриевна,2А33
Шепель Артем Олегович,2А33
Китков Игорь Игоревич,2А34
Гагкаева Мария Аркадьевна,2А34
Варвус Иван Михайлович,2А49
Маркин Александр Андреевич,2А49
Тетерев Ярослав Тарасович,2А52
Джафаров Башир Темирханович,2А52
Петрова Полина Юрьевна,2А52
Пусь Юлия Владимировна,2А54
Прокофьев Евгений Андреевич,2А54
Захаров Александр Сергеевич,2А54
Найданова Лидия Вадимовна,2А55
Нам Ирина Николаевна,2А55|2А54
Танкиева Фатима Умат-Гиреевна,2А55
"""
        vacations_csv = """\
Боброва Эльвира Анатольевна,01.10.2025,12.10.2025
Тетерев Ярослав Тарасович,29.09.2025,12.10.2025
Исаев Абсалам Темурланович,06.10.2025,19.10.2025
Гагкаева Мария Аркадьевна,01.10.2025,14.10.2025
Гагкаева Мария Аркадьевна,20.10.2025,20.10.2025
Гагкаева Мария Аркадьевна,27.10.2025,27.10.2025
Егуренкова Виктория Сергеевна,06.10.2025,19.10.2025
"""
        cabins_text = "2А03, 2А04, 2А05, 2А06, 2А07, 2А08, 2А33, 2А34, 2А49, 2А52, 2А54, 2А55"
        holidays_text = ""

        self.txt_doctors.delete("1.0", "end"); self.txt_doctors.insert("1.0", doctors_csv)
        self.txt_priorities.delete("1.0", "end"); self.txt_priorities.insert("1.0", priorities_csv)
        self.txt_vacations.delete("1.0", "end"); self.txt_vacations.insert("1.0", vacations_csv)
        self.txt_cabins.delete("1.0", "end"); self.txt_cabins.insert("1.0", cabins_text)
        self.txt_holidays.delete("1.0", "end"); self.txt_holidays.insert("1.0", holidays_text)

    def recompute(self):
        try:
            self.status_var.set("Пересчёт...")
            self.update_idletasks()

            year = int(self.year_var.get())
            month = int(self.month_var.get())

            doctors = parse_doctors_csv(self.txt_doctors.get("1.0", "end"))
            pr = parse_priorities_csv(self.txt_priorities.get("1.0", "end"))
            for d in doctors:
                d.priorities = pr.get(d.name, [])

            vacations = parse_vacations_csv(self.txt_vacations.get("1.0", "end"), year, month)
            cabins = parse_cabins(self.txt_cabins.get("1.0", "end"))
            holidays = parse_holidays(self.txt_holidays.get("1.0", "end"), year, month)

            if not cabins:
                raise ValueError("Список кабинетов пуст.")
            if not doctors:
                raise ValueError("Список врачей пуст.")

            days, required, sched, slot_assign, dev = schedule_month(
                doctors=doctors,
                vacations=vacations,
                cabins=cabins,
                year=year,
                month=month,
                holidays=holidays,
            )

            self.doctors = doctors
            self.vacations = vacations
            self.cabins = cabins
            self.holidays = holidays
            self.days = days
            self.required = required
            self.sched = sched
            self.slot_assign = slot_assign

            self._render_table()

            free_slots = sum(1 for d in days for sh, cab, who in slot_assign[d] if who == "свободно")
            self.status_var.set(f"Готово. Свободных слотов: {free_slots}")

        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            self.status_var.set("Ошибка")

    def _render_table(self):
        # clear
        self.tree.delete(*self.tree.get_children())

        cols = ["Врач"] + [str(d.day) for d in self.days] + ["Смен", "Комментарий"]
        self.tree["columns"] = cols

        self.tree.heading("Врач", text="Врач")
        self.tree.column("Врач", width=260, anchor="w")

        for d in self.days:
            k = str(d.day)
            self.tree.heading(k, text=k)
            self.tree.column(k, width=95, anchor="center")

        self.tree.heading("Смен", text="Смен")
        self.tree.column("Смен", width=60, anchor="center")
        self.tree.heading("Комментарий", text="Комментарий")
        self.tree.column("Комментарий", width=190, anchor="w")

        for doc in self.doctors:
            row = [doc.name]
            shift_count = 0
            for day in self.days:
                code, cab = self.sched[doc.name][day]
                if code in ("у", "в", "р"):
                    row.append(f"{code} ({cab})")
                    shift_count += 1
                else:
                    row.append(code)
            row.append(str(shift_count))
            rate = "1.0" if doc.fte >= 0.99 else "0.5"
            row.append(f"ставка {rate}, норма {self.required[doc.name]}")
            self.tree.insert("", "end", values=row)

    def export(self):
        if not self.sched:
            messagebox.showwarning("Нет данных", "Сначала нажмите Recompute.")
            return

        default_name = f"график_{self.year_var.get()}_{self.month_var.get():02d}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Сохранить XLSX",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            export_xlsx(
                path=path,
                doctors=self.doctors,
                days=self.days,
                sched=self.sched,
                required=self.required,
                slot_assign=self.slot_assign,
                cabins=self.cabins,
            )
            messagebox.showinfo("Готово", f"Сохранено:\n{path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))


if __name__ == "__main__":
    app = SchedulerApp()
    app.mainloop()
