#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tkinter UI: physician shift scheduling with OR-Tools CP-SAT + XLSX export
(includes wish_list.xlsx support + per-doctor shift preference + morning/evening mix constraint).

Install:
  pip install ortools openpyxl pandas

Run:
  python scheduler_tk_ortools_wishlist_shiftpref.py
"""

import re
import datetime as dt
from dataclasses import dataclass, field
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Set, Optional
from io import BytesIO

import pandas as pd
from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, messagebox, filedialog


# ----------------------------
# Data model
# ----------------------------

@dataclass
class Doctor:
    name: str
    fte: float = 1.0
    priorities: List[str] = field(default_factory=list)

@dataclass
class Vacation:
    name: str
    start: dt.date
    end: dt.date  # inclusive


# ----------------------------
# Helpers / parsing
# ----------------------------

def parse_date(s: str, year: int, month: int) -> dt.date:
    s = str(s).strip()
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        d, m, y = map(int, s.split("."))
        return dt.date(y, m, d)
    if re.match(r"^\d{2}\.\d{2}$", s):
        d, m = map(int, s.split("."))
        return dt.date(year, m, d)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        y, m, d = map(int, s.split("-"))
        return dt.date(y, m, d)
    raise ValueError(f"Bad date: {s}")

def daterange(a: dt.date, b: dt.date):
    cur = a
    while cur <= b:
        yield cur
        cur += dt.timedelta(days=1)

def is_weekend(d: dt.date) -> bool:
    return d.weekday() >= 5

def all_days_in_month(year: int, month: int) -> List[dt.date]:
    days = []
    d = dt.date(year, month, 1)
    while d.month == month:
        days.append(d)
        d += dt.timedelta(days=1)
    return days

def working_days_in_month(year: int, month: int) -> List[dt.date]:
    return [d for d in all_days_in_month(year, month) if not is_weekend(d)]

def parse_doctors_csv(text: str) -> List[Doctor]:
    out: List[Doctor] = []
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if not parts or not parts[0]:
            continue
        name = parts[0]
        fte = 1.0
        if len(parts) >= 2 and parts[1]:
            try:
                fte = float(parts[1])
            except ValueError:
                raise ValueError(f"Ошибка в списке врачей: ставка должна быть 1 или 0.5. Получено '{parts[1]}' в строке '{line}'")
        out.append(Doctor(name=name, fte=fte))
    return out

def parse_priorities_csv(text: str) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        if len(parts) < 2:
            continue
        name = parts[0]
        cabs = [c.strip() for c in parts[1].split("|") if c.strip()]
        out[name] = cabs
    return out

def parse_vacations_csv(text: str, year: int, month: int) -> List[Vacation]:
    out: List[Vacation] = []
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        name = parts[0]
        start = parse_date(parts[1], year, month)
        end = parse_date(parts[2], year, month)
        out.append(Vacation(name=name, start=start, end=end))
    return out

def parse_cabins(text: str) -> List[str]:
    items = re.split(r"[,\s]+", (text or "").strip())
    return [i for i in (x.strip() for x in items) if i]

def parse_holidays(text: str, year: int, month: int) -> Set[dt.date]:
    """
    Tokens separated by comma/space/newline, each token can be:
      - DD.MM.YYYY
      - DD.MM (year from UI)
      - YYYY-MM-DD
    """
    out=set()
    if not (text or "").strip():
        return out
    for tok in re.split(r"[,\s]+", text.strip()):
        if not tok.strip():
            continue
        out.add(parse_date(tok.strip(), year, month))
    return out

def parse_yes_list(text: str) -> Set[str]:
    out=set()
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        out.add(line.split(",")[0].strip())
    return out

_LATIN_TO_CYR = str.maketrans({
    "A":"А","B":"В","C":"С","E":"Е","H":"Н","K":"К","M":"М","O":"О","P":"Р","T":"Т","X":"Х","Y":"У",
    "a":"а","b":"в","c":"с","e":"е","h":"н","k":"к","m":"м","o":"о","p":"р","t":"т","x":"х","y":"у",
})

def _norm_cyr(s: str) -> str:
    return (s or "").translate(_LATIN_TO_CYR).replace("Ё","Е").replace("ё","е").strip()

def parse_shift_pref_csv(text: str) -> Dict[str, Optional[str]]:
    """
    CSV lines: ФИО,у|в|нет
    """
    out: Dict[str, Optional[str]] = {}
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        name = parts[0]
        pref = _norm_cyr(parts[1].strip().lower()) if len(parts) > 1 else ""
        if pref in ("у", "утро", "morning"):
            out[name] = "у"
        elif pref in ("в", "вечер", "evening"):
            out[name] = "в"
        elif pref in ("нет", "none", "-", ""):
            out[name] = None
        else:
            raise ValueError(f"Неверное предпочтение смены для '{name}': '{parts[1] if len(parts)>1 else ''}'. Используйте у/в/нет")
    return out

def build_vac_map(vacs: List[Vacation]) -> Dict[str, Set[dt.date]]:
    m = defaultdict(set)
    for v in vacs:
        for day in daterange(v.start, v.end):
            m[v.name].add(day)
    return m


# ----------------------------
# Wish-list (optional)
# ----------------------------

def doctor_key_fullname(full_name: str) -> str:
    parts = _norm_cyr(full_name).split()
    if not parts:
        return ""
    sur = parts[0].upper()
    ini = ""
    if len(parts) >= 2 and parts[1]:
        ini += parts[1][0].upper() + "."
    if len(parts) >= 3 and parts[2]:
        ini += parts[2][0].upper() + "."
    return f"{sur} {ini}".strip()

def doctor_key_initials(initials_name: str) -> str:
    s = _norm_cyr(initials_name).upper()
    s = re.sub(r"\s+", " ", s)
    m = re.match(r"^([А-ЯA-Z\-]+)\s+([А-ЯA-Z])\.\s*([А-ЯA-Z])\.$", s)
    if m:
        return f"{m.group(1)} {m.group(2)}.{m.group(3)}."
    return s

def _extract_dates(text: str, year: int, month: int) -> Set[dt.date]:
    out=set()
    if not text or not isinstance(text, str):
        return out
    t = _norm_cyr(text)
    for m in re.finditer(r"(\d{1,2}\.\d{1,2})(?:\.\d{4})?\s*[-–—]\s*(\d{1,2}\.\d{1,2})(?:\.\d{4})?", t):
        a = parse_date(m.group(1), year, month)
        b = parse_date(m.group(2), year, month)
        for d in daterange(a,b):
            if d.year == year and d.month == month:
                out.add(d)
    for m in re.finditer(r"(\d{1,2}\.\d{1,2}(?:\.\d{4})?)", t):
        try:
            d = parse_date(m.group(1), year, month)
            if d.year == year and d.month == month:
                out.add(d)
        except Exception:
            pass
    for m in re.finditer(r"(\d{4}-\d{2}-\d{2})", t):
        try:
            d = parse_date(m.group(1), year, month)
            if d.year == year and d.month == month:
                out.add(d)
        except Exception:
            pass
    return out

def parse_wishlist_xlsx(file_bytes: bytes, doctors: List[Doctor], year: int, month: int) -> Dict:
    df = pd.read_excel(BytesIO(file_bytes))
    full_by_key = {doctor_key_fullname(d.name): d.name for d in doctors}
    out = {
        "extra_ok_yes": set(),
        "extra_ok_no": set(),
        "add_vac": [],
        "date_off_hard": defaultdict(set),
        "weekday_shift_pref": defaultdict(dict),
        "evenodd_shift_pref": {},
        "priority_override": {},
        "pref_shift": {},
    }
    col_name = next((c for c in df.columns if "Фам" in str(c)), df.columns[0])
    col_sched = next((c for c in df.columns if "график" in str(c).lower()), None)
    col_vac = next((c for c in df.columns if "Даты" in str(c) and "отпуск" in str(c).lower()), None)
    col_extra = next((c for c in df.columns if "Дополнитель" in str(c)), None)
    col_prio = next((c for c in df.columns if "Приоритет" in str(c)), None)
    col_other = next((c for c in df.columns if "Другие" in str(c)), None)

    weekday_words = {
        "понедель":0, "вторник":1, "вторн":1, "сред":2, "четвер":3, "пятниц":4, "суббот":5, "воскрес":6
    }

    for _, row in df.iterrows():
        raw_name = str(row.get(col_name, "")).strip()
        if not raw_name or raw_name.lower() == "nan":
            continue
        key = doctor_key_initials(raw_name)
        full = full_by_key.get(key)
        if not full:
            continue

        if col_prio is not None:
            cab = _norm_cyr(str(row.get(col_prio, "")).strip())
            if cab and cab.lower() != "nan":
                out["priority_override"][full] = cab

        if col_extra is not None:
            extra = _norm_cyr(str(row.get(col_extra, "")).strip()).lower()
            if "не хочу" in extra or "не могу" in extra:
                out["extra_ok_no"].add(full)
            elif "могу" in extra or "подзаработ" in extra or "доп" in extra:
                out["extra_ok_yes"].add(full)

        if col_vac is not None:
            dates = _extract_dates(str(row.get(col_vac, "")), year, month)
            for d in sorted(dates):
                out["add_vac"].append(Vacation(name=full, start=d, end=d))

        if col_sched is not None:
            sched_text = _norm_cyr(str(row.get(col_sched, ""))).lower()
            if "четн" in sched_text and "нечетн" in sched_text:
                m1 = re.search(r"четн\w*\s*(утр|вечер)", sched_text)
                m2 = re.search(r"нечетн\w*\s*(утр|вечер)", sched_text)
                if m1 and m2:
                    even = "у" if m1.group(1).startswith("утр") else "в"
                    odd = "у" if m2.group(1).startswith("утр") else "в"
                    out["evenodd_shift_pref"][full] = {"even": even, "odd": odd}

        if col_other is not None:
            other = _norm_cyr(str(row.get(col_other, ""))).lower()
            if other and other != "nan":
                dates = _extract_dates(other, year, month)
                if ("выходн" in other) or ("не став" in other) or ("не работать" in other):
                    for d in dates:
                        out["date_off_hard"][full].add(d)
                for ww, wdi in weekday_words.items():
                    if ww in other:
                        pref = None
                        if ("1" in other) or ("перв" in other):
                            pref = "у"
                        if ("2" in other) or ("вторую смен" in other):
                            pref = "в"
                        if pref in ("у","в"):
                            out["weekday_shift_pref"][full][wdi] = pref

    return out


# ----------------------------
# Priority collision rule
# ----------------------------

def apply_priority_collision_rule(doctors: List[Doctor]) -> Tuple[List[Doctor], List[Tuple[str,str]]]:
    top_map = defaultdict(list)
    for d in doctors:
        if d.priorities:
            top_map[d.priorities[0]].append(d)
    removed = []
    for cabin, ds in top_map.items():
        if len(ds) >= 3:
            half = [d for d in ds if d.fte < 0.99]
            if half:
                pick = sorted(half, key=lambda x: x.name)[-1]
            else:
                pick = sorted(ds, key=lambda x: x.name)[-1]
            removed.append((pick.name, cabin))
            pick.priorities = []
    return doctors, removed


# ----------------------------
# Calendar / norms / slots
# ----------------------------

def required_norm(doc: Doctor, workdays: List[dt.date], vac_days: Set[dt.date]) -> int:
    base = 22 if doc.fte >= 0.99 else 11
    vac_wd = sum(1 for d in workdays if d in vac_days)
    return max(0, base - vac_wd)

def build_slots(days: List[dt.date], cabins: List[str], holidays: Set[dt.date]):
    slots = []
    for di, day in enumerate(days):
        wkend = is_weekend(day) or (day in holidays)
        shifts = ['р'] if wkend else ['у','в']
        for sh in shifts:
            for cab in cabins:
                slots.append((di, day, sh, cab))
    return slots


# ----------------------------
# Solver
# ----------------------------

def solve_with_cpsat(
    doctors: List[Doctor],
    vacations: List[Vacation],
    cabins: List[str],
    year: int,
    month: int,
    holidays: Set[dt.date],
    extra_ok: Set[str],
    wish: Optional[Dict] = None,
    shift_pref: Optional[Dict[str, Optional[str]]] = None,
    pref_weight: int = 3,
    enforce_morning_evening_mix: bool = True,
    extra_max: int = 6,
    time_limit_s: int = 30,
    free_label: str = "свободно",
    num_workers: int = 8,
):
    days = all_days_in_month(year, month)
    workdays = working_days_in_month(year, month)
    vac_map = build_vac_map(vacations)
    norm = {d.name: required_norm(d, workdays, vac_map.get(d.name,set())) for d in doctors}

    doctors, removed = apply_priority_collision_rule(doctors)
    doc_names = [d.name for d in doctors]

    slots = build_slots(days, cabins, holidays)
    slot_idx = {(di, sh, cab): si for si,(di,_,sh,cab) in enumerate(slots)}

    model = cp_model.CpModel()

    x = {}
    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, name in enumerate(doc_names):
            if day in vac_map.get(name, set()):
                continue
            x[(si,dj)] = model.NewBoolVar(f"x_s{si}_d{dj}")

    free = [model.NewBoolVar(f"free_{si}") for si in range(len(slots))]

    for si in range(len(slots)):
        vars_in = [x[(si,dj)] for dj in range(len(doc_names)) if (si,dj) in x]
        model.Add(sum(vars_in) + free[si] == 1)

    work = {}
    for di, day in enumerate(days):
        wkend = is_weekend(day) or (day in holidays)
        shifts = ['р'] if wkend else ['у','в']
        for dj in range(len(doc_names)):
            w = model.NewBoolVar(f"work_{di}_{dj}")
            work[(di,dj)] = w
            vars_in = []
            for cab in cabins:
                for sh in shifts:
                    si = slot_idx[(di, sh, cab)]
                    if (si,dj) in x:
                        vars_in.append(x[(si,dj)])
            if vars_in:
                model.Add(sum(vars_in) == w)  # implies <=1 shift/day per doctor
            else:
                model.Add(w == 0)

    if wish and wish.get("date_off_hard"):
        for dj, name in enumerate(doc_names):
            for di, day in enumerate(days):
                if day in wish["date_off_hard"].get(name, set()):
                    model.Add(work[(di,dj)] == 0)

    for dj in range(len(doc_names)):
        for start in range(0, len(days)-6+1):
            model.Add(sum(work[(di,dj)] for di in range(start, start+6)) <= 5)

    if enforce_morning_evening_mix:
        for dj, name in enumerate(doc_names):
            weekday_m = []
            weekday_e = []
            for di, day in enumerate(days):
                if is_weekend(day) or (day in holidays):
                    continue
                for cab in cabins:
                    si_m = slot_idx[(di, 'у', cab)]
                    si_e = slot_idx[(di, 'в', cab)]
                    if (si_m, dj) in x:
                        weekday_m.append(x[(si_m, dj)])
                    if (si_e, dj) in x:
                        weekday_e.append(x[(si_e, dj)])
            if not weekday_m and not weekday_e:
                continue
            m_cnt = sum(weekday_m) if weekday_m else 0
            e_cnt = sum(weekday_e) if weekday_e else 0
            tot = m_cnt + e_cnt
            has_two = model.NewBoolVar(f"has_two_weekday_{dj}")
            model.Add(tot >= 2).OnlyEnforceIf(has_two)
            model.Add(tot <= 1).OnlyEnforceIf(has_two.Not())
            model.Add(m_cnt >= 1).OnlyEnforceIf(has_two)
            model.Add(e_cnt >= 1).OnlyEnforceIf(has_two)

    for dj, name in enumerate(doc_names):
        total_work = sum(work[(di,dj)] for di in range(len(days)))
        if name in extra_ok:
            model.Add(total_work >= norm[name])
            model.Add(total_work <= norm[name] + extra_max)
        else:
            model.Add(total_work == norm[name])

    pri_weight_full = 8
    pri_weight_half = 3
    fill_weight = 20

    obj_terms = []
    for si in range(len(slots)):
        obj_terms.append(fill_weight * (1 - free[si]))

    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, doc in enumerate(doctors):
            if (si,dj) not in x:
                continue
            if cab in doc.priorities:
                w = pri_weight_half if doc.fte < 0.99 else pri_weight_full
                obj_terms.append(w * x[(si,dj)])

    wish_weight = 2
    if wish and wish.get("pref_shift"):
        for (di, dj), pref in wish["pref_shift"].items():
            if pref not in ("у", "в"):
                continue
            day = days[di]
            if is_weekend(day) or (day in holidays):
                continue
            vars_pref = []
            for cab in cabins:
                si = slot_idx[(di, pref, cab)]
                if (si, dj) in x:
                    vars_pref.append(x[(si, dj)])
            if vars_pref:
                obj_terms.append(wish_weight * sum(vars_pref))

    if shift_pref:
        for dj, doc in enumerate(doctors):
            pref = shift_pref.get(doc.name)
            if pref not in ('у', 'в'):
                continue
            for di, day in enumerate(days):
                if is_weekend(day) or (day in holidays):
                    continue
                vars_match = []
                for cab in cabins:
                    si = slot_idx[(di, pref, cab)]
                    if (si, dj) in x:
                        vars_match.append(x[(si, dj)])
                if vars_match:
                    obj_terms.append(int(pref_weight) * sum(vars_match))

    model.Maximize(sum(obj_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.num_search_workers = int(max(1, num_workers))

    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("CP-SAT не нашёл допустимое решение (увеличьте time_limit или разрешите доп.смены extra_ok).")

    sched: Dict[str, Dict[dt.date, Tuple[str, str]]] = {n:{} for n in doc_names}
    for name in doc_names:
        for day in days:
            if day in vac_map.get(name,set()):
                sched[name][day] = ("от","")
            else:
                sched[name][day] = ("-","")

    slot_assign: Dict[dt.date, List[Tuple[str,str,str]]] = {d: [] for d in days}
    for si, (di, day, sh, cab) in enumerate(slots):
        who = free_label
        for dj, name in enumerate(doc_names):
            if (si,dj) in x and solver.Value(x[(si,dj)]) == 1:
                who = name
                break
        slot_assign[day].append((sh,cab,who))
        if who != free_label:
            sched[who][day] = (sh,cab)

    deviation = {}
    for name in doc_names:
        fact = sum(1 for day in days if sched[name][day][0] in ("у","в","р"))
        deviation[name] = fact - norm[name]

    meta = {
        "removed_priorities": removed,
        "objective": solver.ObjectiveValue(),
        "status": "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE",
    }
    return days, norm, sched, slot_assign, deviation, meta, doctors


# ----------------------------
# XLSX export
# ----------------------------

def export_xlsx_bytes(doctors: List[Doctor],
                      days: List[dt.date],
                      sched: Dict[str, Dict[dt.date, Tuple[str, str]]],
                      norm: Dict[str, int],
                      slot_assign: Dict[dt.date, List[Tuple[str, str, str]]],
                      cabins: List[str],
                      free_label: str = "свободно") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "График врачей"

    fill_morning = PatternFill("solid", fgColor="A7D8FF")  # голубой
    fill_evening = PatternFill("solid", fgColor="FFB6C1")  # розовый
    fill_weekend = PatternFill("solid", fgColor="A7F3A7")  # зелёный
    fill_vac = PatternFill("solid", fgColor="C0C0C0")      # серый
    fill_off = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Врач").font = header_font
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=d.day)
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 11

    last_col = 2 + len(days)
    ws.cell(row=1, column=last_col, value="Смен").font = header_font
    ws.cell(row=1, column=last_col + 1, value="Комментарий").font = header_font

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(last_col)].width = 8
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 28

    for r, doc in enumerate(doctors, start=2):
        ws.cell(row=r, column=1, value=doc.name).border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)

        shift_count = 0
        for i, day in enumerate(days, start=2):
            code, cab = sched[doc.name][day]
            txt = code
            if code in ("у", "в", "р"):
                txt = f"{code} ({cab})"
                shift_count += 1
            cell = ws.cell(row=r, column=i, value=txt)
            cell.alignment = center
            cell.border = border
            if code == "у":
                cell.fill = fill_morning
            elif code == "в":
                cell.fill = fill_evening
            elif code == "р":
                cell.fill = fill_weekend
            elif code == "от":
                cell.fill = fill_vac
            else:
                cell.fill = fill_off

        ws.cell(row=r, column=last_col, value=shift_count).alignment = center
        ws.cell(row=r, column=last_col).border = border

        rate = "1.0" if doc.fte >= 0.99 else "0.5"
        ws.cell(row=r, column=last_col + 1, value=f"ставка {rate}, норма {norm[doc.name]}").border = border
        ws.cell(row=r, column=last_col + 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("Сводка по врачам")
    ws2.append(["Врач", "Ставка", "Норма смен", "Факт смен", "Отклонение", "Приоритеты"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        ws2.append([doc.name, doc.fte, norm[doc.name], fact, fact - norm[doc.name], ", ".join(doc.priorities)])
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=6):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["F"].width = 26

    # Sheet 3
    ws3 = wb.create_sheet("Загрузка кабинетов")
    ws3.append(["Дата", "День", "Смена", "Кабинет", "Врач"])
    for c in ws3[1]:
        c.font = header_font
        c.border = border
        c.alignment = center
    day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for d in days:
        for sh, cab, who in slot_assign[d]:
            ws3.append([d.isoformat(), day_name[d.weekday()], sh, cab, who])
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 6
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 34

    # Sheet 4
    ws4 = wb.create_sheet("Общая статистика")
    ws4.append(["Показатель", "Значение"])
    ws4["A1"].font = header_font
    ws4["B1"].font = header_font
    total_slots = sum(len(slot_assign[d]) for d in days)
    free_slots = sum(1 for d in days for sh, cab, who in slot_assign[d] if who == free_label)
    ws4.append(["Месяц", f"{days[0].strftime('%B')} {days[0].year}"])
    ws4.append(["Кабинетов", len(cabins)])
    ws4.append(["Всего слотов (кабинет-смена)", total_slots])
    ws4.append(["Свободно (не заполнено)", free_slots])
    ws4.append(["Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0])
    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18

    # Sheet 5
    ws5 = wb.create_sheet("Обоснование отклонений")
    ws5.append(["Врач", "Норма", "Факт", "Отклонение", "Пояснение"])
    for c in ws5[1]:
        c.font = header_font
        c.border = border
        c.alignment = center
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ("у", "в", "р"))
        dev = fact - norm[doc.name]
        vac_count = sum(1 for d in days if sched[doc.name][d][0] == "от")
        expl = []
        if vac_count:
            expl.append(f"отпуск: {vac_count} дн.")
        if dev > 0:
            expl.append(f"доп. смены: {dev}")
        elif dev < 0:
            expl.append(f"не добрали: {-dev}")
        else:
            expl.append("норма выполнена")
        ws5.append([doc.name, norm[doc.name], fact, dev, "; ".join(expl)])
    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["E"].width = 42

    # Sheet 6
    ws6 = wb.create_sheet("Кабинетов на врача")
    ws6.append(["Врач", "Кабинет", "Кол-во смен"])
    for c in ws6[1]:
        c.font = header_font
        c.border = border
        c.alignment = center
    for doc in doctors:
        cab_counter = Counter()
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ("у", "в", "р"):
                cab_counter[cab] += 1
        for cab, cnt in cab_counter.most_common():
            ws6.append([doc.name, cab, cnt])
    for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws6.column_dimensions["A"].width = 34
    ws6.column_dimensions["B"].width = 10
    ws6.column_dimensions["C"].width = 14

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Tkinter UI
# ----------------------------

DEFAULT_DOCTORS = """\
Боброва Эльвира Анатольевна,1
Буданова Татьяна Владимировна,0.5
Варвус Иван Михайлович,1
Гагкаева Мария Аркадьевна,1
Гущина Юлия Викторовна,1
Джафаров Башир Темирханович,0.5
Догузова Александра Гочевна,1
Евдокимова Жанна Владимировна,1
Егуренкова Виктория Сергеевна,1
Захаров Александр Сергеевич,1
Исаев Абсалам Темурланович,1
Китков Игорь Игоревич,1
Кокорина Серафима Евгеньевна,1
Королькова Александрова Викторовна,0.5
Короткова Полина Александровна,1
Лукьяненко Владимир Александрович,1
Маркин Александр Андреевич,1
Марфутов Василий Васильевич,1
Медведева Мария Вячеславовна,1
Найданова Лидия Вадимовна,1
Нам Ирина Николаевна,1
Петрова Полина Юрьевна,1
Пономарева Юлия Константиновна,1
Прокофьев Евгений Андреевич,1
Пусь Юлия Владимировна,1
Танкиева Фатима Умат-Гиреевна,1
Тетерев Ярослав Тарасович,1
Федорова Оксана Сергеевна,1
Харитонова Наталья Игоревна,1
Шепель Артем Олегович,1
Юдин Алексей Александрович,1
Юдина Виктория Дмитриевна,1
"""

DEFAULT_PRIORITIES = """\
Пономарева Юлия Константиновна,2А03
Лукьяненко Владимир Александрович,2А03
Королькова Александрова Викторовна,2А03
Харитонова Наталья Игоревна,2А04
Исаев Абсалам Темурланович,2А04
Короткова Полина Александровна,2А04
Гущина Юлия Викторовна,2А05
Федорова Оксана Сергеевна,2А05
Кокорина Серафима Евгеньевна,2А05
Егуренкова Виктория Сергеевна,2А06
Евдокимова Жанна Владимировна,2А06
Буданова Татьяна Владимировна,2А06
Боброва Эльвира Анатольевна,2А07
Марфутов Василий Васильевич,2А07
Догузова Александра Гочевна,2А07
Медведева Мария Вячеславовна,2А08
Юдин Алексей Александрович,2А08
Юдина Виктория Дмитриевна,2А33
Шепель Артем Олегович,2А33
Китков Игорь Игоревич,2А34
Гагкаева Мария Аркадьевна,2А34
Варвус Иван Михайлович,2А49
Маркин Александр Андреевич,2А49
Тетерев Ярослав Тарасович,2А52
Джафаров Башир Темирханович,2А52
Петрова Полина Юрьевна,2А52
Пусь Юлия Владимировна,2А54
Прокофьев Евгений Андреевич,2А54
Захаров Александр Сергеевич,2А54
Найданова Лидия Вадимовна,2А55
Нам Ирина Николаевна,2А55|2А54
Танкиева Фатима Умат-Гиреевна,2А55
"""

DEFAULT_VAC = """\
Боброва Эльвира Анатольевна,01.10.2025,12.10.2025
Тетерев Ярослав Тарасович,29.09.2025,12.10.2025
Исаев Абсалам Темурланович,06.10.2025,19.10.2025
Гагкаева Мария Аркадьевна,01.10.2025,14.10.2025
Гагкаева Мария Аркадьевна,20.10.2025,20.10.2025
Гагкаева Мария Аркадьевна,27.10.2025,27.10.2025
Егуренкова Виктория Сергеевна,06.10.2025,19.10.2025
"""

DEFAULT_CABINS = "2А03, 2А04, 2А05, 2А06, 2А07, 2А08, 2А33, 2А34, 2А49, 2А52, 2А54, 2А55"


class SchedulerTkApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("График смен врачей — OR-Tools CP-SAT (Tkinter)")
        self.root.geometry("1400x850")

        self.wish_path: Optional[str] = None
        self.last_result = None  # (days, norm, sched, slot_assign, meta, doctors, cabins)

        self._build_ui()

    def _build_ui(self):
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        main = ttk.Frame(self.root, padding=8)
        main.grid(row=0, column=0, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        # Top controls
        top = ttk.Frame(main)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for i in range(16):
            top.columnconfigure(i, weight=0)

        ttk.Label(top, text="Год:").grid(row=0, column=0, sticky="w")
        self.var_year = tk.IntVar(value=2025)
        ttk.Entry(top, textvariable=self.var_year, width=6).grid(row=0, column=1, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Месяц:").grid(row=0, column=2, sticky="w")
        self.var_month = tk.IntVar(value=10)
        ttk.Entry(top, textvariable=self.var_month, width=4).grid(row=0, column=3, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Time limit (сек):").grid(row=0, column=4, sticky="w")
        self.var_time = tk.IntVar(value=30)
        ttk.Entry(top, textvariable=self.var_time, width=6).grid(row=0, column=5, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Extra max:").grid(row=0, column=6, sticky="w")
        self.var_extra_max = tk.IntVar(value=6)
        ttk.Entry(top, textvariable=self.var_extra_max, width=4).grid(row=0, column=7, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Вес pref (у/в):").grid(row=0, column=8, sticky="w")
        self.var_pref_weight = tk.IntVar(value=3)
        ttk.Entry(top, textvariable=self.var_pref_weight, width=4).grid(row=0, column=9, sticky="w", padx=(4, 12))

        self.var_enforce_mix = tk.BooleanVar(value=True)
        ttk.Checkbutton(top, text="Требовать смесь утро/вечер (если ≥2 будних смен)", variable=self.var_enforce_mix)\
            .grid(row=0, column=10, sticky="w", padx=(0, 12))

        self.var_use_wish = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text="Использовать wish_list.xlsx", variable=self.var_use_wish)\
            .grid(row=0, column=11, sticky="w")

        ttk.Button(top, text="Выбрать wish_list.xlsx", command=self.pick_wishlist)\
            .grid(row=0, column=12, sticky="w", padx=(6, 12))

        ttk.Button(top, text="Recompute (CP-SAT)", command=self.on_recompute)\
            .grid(row=0, column=13, sticky="w", padx=(0, 12))

        ttk.Button(top, text="Export XLSX...", command=self.on_export_xlsx)\
            .grid(row=0, column=14, sticky="w")

        # Notebook
        nb = ttk.Notebook(main)
        nb.grid(row=1, column=0, sticky="nsew")
        self.nb = nb

        self.tab_inputs = ttk.Frame(nb)
        self.tab_output = ttk.Frame(nb)

        nb.add(self.tab_inputs, text="Ввод")
        nb.add(self.tab_output, text="Результат")

        self._build_inputs_tab()
        self._build_output_tab()

        # Status bar
        self.status = tk.StringVar(value="Готово.")
        ttk.Label(main, textvariable=self.status, anchor="w").grid(row=2, column=0, sticky="ew", pady=(8, 0))

    def _make_scrolled_text(self, parent, height=10, width=60):
        frame = ttk.Frame(parent)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        txt = tk.Text(frame, height=height, width=width, wrap="none")
        y = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        x = ttk.Scrollbar(frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=y.set, xscrollcommand=x.set)
        txt.grid(row=0, column=0, sticky="nsew")
        y.grid(row=0, column=1, sticky="ns")
        x.grid(row=1, column=0, sticky="ew")
        return frame, txt

    def _build_inputs_tab(self):
        t = self.tab_inputs
        t.rowconfigure(0, weight=1)
        t.columnconfigure(0, weight=1)

        container = ttk.Frame(t, padding=8)
        container.grid(row=0, column=0, sticky="nsew")
        for c in range(2):
            container.columnconfigure(c, weight=1)
        for r in range(4):
            container.rowconfigure(r, weight=1)

        # Doctors
        lf1 = ttk.LabelFrame(container, text="Врачи (CSV: ФИО,ставка 1/0.5)")
        lf1.grid(row=0, column=0, sticky="nsew", padx=(0,8), pady=(0,8))
        f, self.txt_doctors = self._make_scrolled_text(lf1, height=12)
        f.grid(row=0, column=0, sticky="nsew")
        lf1.rowconfigure(0, weight=1); lf1.columnconfigure(0, weight=1)
        self.txt_doctors.insert("1.0", DEFAULT_DOCTORS)

        # Priorities
        lf2 = ttk.LabelFrame(container, text="Приоритеты (CSV: ФИО,кабинет1|кабинет2|...)")
        lf2.grid(row=0, column=1, sticky="nsew", pady=(0,8))
        f, self.txt_prio = self._make_scrolled_text(lf2, height=12)
        f.grid(row=0, column=0, sticky="nsew")
        lf2.rowconfigure(0, weight=1); lf2.columnconfigure(0, weight=1)
        self.txt_prio.insert("1.0", DEFAULT_PRIORITIES)

        # Vacations
        lf3 = ttk.LabelFrame(container, text="Отпуска (CSV: ФИО,DD.MM.YYYY,DD.MM.YYYY)")
        lf3.grid(row=1, column=0, sticky="nsew", padx=(0,8), pady=(0,8))
        f, self.txt_vac = self._make_scrolled_text(lf3, height=10)
        f.grid(row=0, column=0, sticky="nsew")
        lf3.rowconfigure(0, weight=1); lf3.columnconfigure(0, weight=1)
        self.txt_vac.insert("1.0", DEFAULT_VAC)

        # Cabins
        lf4 = ttk.LabelFrame(container, text="Кабинеты (через запятую/пробел)")
        lf4.grid(row=1, column=1, sticky="nsew", pady=(0,8))
        f, self.txt_cabins = self._make_scrolled_text(lf4, height=5)
        f.grid(row=0, column=0, sticky="nsew")
        lf4.rowconfigure(0, weight=1); lf4.columnconfigure(0, weight=1)
        self.txt_cabins.insert("1.0", DEFAULT_CABINS)

        # Holidays
        lf5 = ttk.LabelFrame(container, text="Праздники/доп. выходные (DD.MM или DD.MM.YYYY или YYYY-MM-DD; через пробел/запятую)")
        lf5.grid(row=2, column=0, sticky="nsew", padx=(0,8), pady=(0,8))
        f, self.txt_holidays = self._make_scrolled_text(lf5, height=4)
        f.grid(row=0, column=0, sticky="nsew")
        lf5.rowconfigure(0, weight=1); lf5.columnconfigure(0, weight=1)

        # Extra ok
        lf6 = ttk.LabelFrame(container, text="Согласны на доп. смены (по одному ФИО в строке, опц.)")
        lf6.grid(row=2, column=1, sticky="nsew", pady=(0,8))
        f, self.txt_extra_ok = self._make_scrolled_text(lf6, height=4)
        f.grid(row=0, column=0, sticky="nsew")
        lf6.rowconfigure(0, weight=1); lf6.columnconfigure(0, weight=1)

        # Shift pref
        lf7 = ttk.LabelFrame(container, text="Предпочтение смены (CSV: ФИО,у/в/нет) — мягкое")
        lf7.grid(row=3, column=0, columnspan=2, sticky="nsew")
        f, self.txt_shift_pref = self._make_scrolled_text(lf7, height=5)
        f.grid(row=0, column=0, sticky="nsew")
        lf7.rowconfigure(0, weight=1); lf7.columnconfigure(0, weight=1)

    def _build_output_tab(self):
        t = self.tab_output
        t.rowconfigure(1, weight=1)
        t.columnconfigure(0, weight=1)

        info = ttk.Frame(t, padding=8)
        info.grid(row=0, column=0, sticky="ew")
        info.columnconfigure(1, weight=1)

        ttk.Label(info, text="Статус:").grid(row=0, column=0, sticky="w")
        self.lbl_solution = ttk.Label(info, text="—")
        self.lbl_solution.grid(row=0, column=1, sticky="w")

        ttk.Label(info, text="Снят приоритет (правило 3 врача/кабинет):").grid(row=1, column=0, sticky="w")
        self.lbl_removed = ttk.Label(info, text="—")
        self.lbl_removed.grid(row=1, column=1, sticky="w")

        # Treeview schedule
        frame = ttk.Frame(t, padding=8)
        frame.grid(row=1, column=0, sticky="nsew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(frame, show="headings")
        y = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        x = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y.set, xscrollcommand=x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y.grid(row=0, column=1, sticky="ns")
        x.grid(row=1, column=0, sticky="ew")

    def pick_wishlist(self):
        path = filedialog.askopenfilename(
            title="Выберите wish_list.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.wish_path = path
            self.var_use_wish.set(True)
            self.status.set(f"wish_list: {path}")

    def _get_text(self, txt: tk.Text) -> str:
        return txt.get("1.0", "end-1c")

    def on_recompute(self):
        try:
            self.status.set("Решаю CP-SAT...")
            self.root.update_idletasks()

            year = int(self.var_year.get())
            month = int(self.var_month.get())
            time_limit = int(self.var_time.get())
            extra_max = int(self.var_extra_max.get())
            pref_weight = int(self.var_pref_weight.get())
            enforce_mix = bool(self.var_enforce_mix.get())

            doctors = parse_doctors_csv(self._get_text(self.txt_doctors))
            pr_map = parse_priorities_csv(self._get_text(self.txt_prio))
            for d in doctors:
                d.priorities = pr_map.get(d.name, [])

            vacations = parse_vacations_csv(self._get_text(self.txt_vac), year, month)
            cabins = parse_cabins(self._get_text(self.txt_cabins))
            holidays = parse_holidays(self._get_text(self.txt_holidays), year, month)
            extra_ok = parse_yes_list(self._get_text(self.txt_extra_ok))
            shift_pref = parse_shift_pref_csv(self._get_text(self.txt_shift_pref))

            if not doctors:
                raise ValueError("Список врачей пуст.")
            if not cabins:
                raise ValueError("Список кабинетов пуст.")

            wish = None
            if self.var_use_wish.get():
                if self.wish_path:
                    with open(self.wish_path, "rb") as f:
                        wish_bytes = f.read()
                else:
                    wish_bytes = None
                if wish_bytes:
                    wish = parse_wishlist_xlsx(wish_bytes, doctors, year, month)

                    # Priority overrides
                    for d in doctors:
                        cab = wish.get("priority_override", {}).get(d.name)
                        if cab:
                            if cab in d.priorities:
                                d.priorities.remove(cab)
                            d.priorities.insert(0, cab)

                    # Add extra vacations
                    add_v = wish.get("add_vac", [])
                    if add_v:
                        vacations.extend(add_v)

                    # Update extra_ok by wish flags
                    extra_ok |= set(wish.get("extra_ok_yes", set()))
                    extra_ok -= set(wish.get("extra_ok_no", set()))

                    # Build preferred shift map (di,dj) -> 'у'/'в'
                    days_for_map = all_days_in_month(year, month)
                    doc_names = [d.name for d in doctors]
                    name_to_dj = {n:i for i,n in enumerate(doc_names)}
                    pref_shift = {}
                    evenodd = wish.get("evenodd_shift_pref", {})
                    weekday_pref = wish.get("weekday_shift_pref", {})
                    for di, day in enumerate(days_for_map):
                        if is_weekend(day) or (day in holidays):
                            continue
                        for name in doc_names:
                            dj = name_to_dj[name]
                            pref = None
                            wd = day.weekday()
                            if name in weekday_pref and wd in weekday_pref[name]:
                                pref = weekday_pref[name][wd]
                            elif name in evenodd:
                                pref = evenodd[name]["even"] if (day.day % 2 == 0) else evenodd[name]["odd"]
                            if pref in ("у","в"):
                                pref_shift[(di,dj)] = pref
                    wish["pref_shift"] = pref_shift

            days, norm, sched, slot_assign, deviation, meta, doctors_final = solve_with_cpsat(
                doctors=doctors,
                vacations=vacations,
                cabins=cabins,
                year=year,
                month=month,
                holidays=holidays,
                extra_ok=extra_ok,
                wish=wish,
                shift_pref=shift_pref,
                pref_weight=pref_weight,
                enforce_morning_evening_mix=enforce_mix,
                extra_max=extra_max,
                time_limit_s=time_limit,
            )

            self.last_result = (days, norm, sched, slot_assign, meta, doctors_final, cabins)
            self._render_schedule(days, norm, sched, doctors_final, meta)
            self.status.set("Готово.")
        except Exception as e:
            self.status.set("Ошибка.")
            messagebox.showerror("Ошибка", str(e))

    def _render_schedule(self, days, norm, sched, doctors, meta):
        # columns: Doctor, 1..N, Shifts, Comment
        cols = ["Врач"] + [str(d.day) for d in days] + ["Смен", "Комментарий"]
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c)
            w = 240 if c == "Врач" else (60 if c in ("Смен",) else (160 if c == "Комментарий" else 75))
            self.tree.column(c, width=w, anchor="center")
        # clear
        for item in self.tree.get_children():
            self.tree.delete(item)
        # insert rows
        for doc in doctors:
            vals = [doc.name]
            shift_count = 0
            for d in days:
                code, cab = sched[doc.name][d]
                if code in ("у","в","р"):
                    vals.append(f"{code} ({cab})")
                    shift_count += 1
                else:
                    vals.append(code)
            vals.append(str(shift_count))
            rate = "1.0" if doc.fte >= 0.99 else "0.5"
            vals.append(f"ставка {rate}, норма {norm[doc.name]}")
            self.tree.insert("", "end", values=vals)

        self.lbl_solution.config(text=f"{meta.get('status')} | objective={meta.get('objective'):.0f}")
        removed = meta.get("removed_priorities") or []
        if removed:
            self.lbl_removed.config(text="; ".join([f"{a} (снят {b})" for a,b in removed]))
        else:
            self.lbl_removed.config(text="—")

    def on_export_xlsx(self):
        if not self.last_result:
            messagebox.showinfo("Экспорт", "Сначала нажмите Recompute.")
            return
        days, norm, sched, slot_assign, meta, doctors, cabins = self.last_result
        try:
            bytes_xlsx = export_xlsx_bytes(
                doctors=doctors,
                days=days,
                sched=sched,
                norm=norm,
                slot_assign=slot_assign,
                cabins=cabins,
            )
            path = filedialog.asksaveasfilename(
                title="Сохранить XLSX",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"график_{self.var_year.get()}_{int(self.var_month.get()):02d}_cpsat.xlsx",
            )
            if not path:
                return
            with open(path, "wb") as f:
                f.write(bytes_xlsx)
            messagebox.showinfo("Экспорт", f"Сохранено:\n{path}")
        except Exception as e:
            messagebox.showerror("Экспорт", str(e))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = SchedulerTkApp()
    app.run()
