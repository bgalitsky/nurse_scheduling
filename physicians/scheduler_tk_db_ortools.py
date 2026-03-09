
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tkinter scheduler with:
- OR-Tools CP-SAT solver
- SQLite database for departments and physicians
- Output tabs similar to Excel sheets
- Optional tksheet for per-cell coloring

Install:
  pip install ortools openpyxl pandas
  pip install tksheet   # optional for per-cell colors

Run:
  python scheduler_tk_db_ortools.py
"""

from __future__ import annotations

import os
import re
import sqlite3
import datetime as dt
from dataclasses import dataclass, field
from collections import defaultdict, Counter
from io import BytesIO
from typing import Dict, List, Tuple, Set, Optional

import pandas as pd
from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from tksheet import Sheet
    HAS_TKSHEET = True
except Exception:
    Sheet = None
    HAS_TKSHEET = False


# ----------------------------
# Models / parsing
# ----------------------------

@dataclass
class Doctor:
    name: str
    fte: float = 1.0
    priorities: List[str] = field(default_factory=list)

@dataclass
class Vacation:
    name: str
    start: dt.date
    end: dt.date


def parse_date(s: str, year: int, month: int) -> dt.date:
    s = str(s).strip()
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        d, m, y = map(int, s.split("."))
        return dt.date(y, m, d)
    if re.match(r"^\d{2}\.\d{2}$", s):
        d, m = map(int, s.split("."))
        return dt.date(year, m, d)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        y, m, d = map(int, s.split("-"))
        return dt.date(y, m, d)
    raise ValueError(f"Bad date format: {s}")

def daterange(a: dt.date, b: dt.date):
    cur = a
    while cur <= b:
        yield cur
        cur += dt.timedelta(days=1)

def is_weekend(d: dt.date) -> bool:
    return d.weekday() >= 5

def all_days_in_month(year: int, month: int) -> List[dt.date]:
    days = []
    d = dt.date(year, month, 1)
    while d.month == month:
        days.append(d)
        d += dt.timedelta(days=1)
    return days

def working_days_in_month(year: int, month: int) -> List[dt.date]:
    return [d for d in all_days_in_month(year, month) if not is_weekend(d)]

def parse_doctors_csv(text: str) -> List[Doctor]:
    out = []
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        name = parts[0]
        fte = 1.0
        if len(parts) > 1 and parts[1]:
            try:
                fte = float(parts[1])
            except ValueError:
                raise ValueError(f"Bad FTE '{parts[1]}' in line '{line}'")
        out.append(Doctor(name=name, fte=fte))
    return out

def parse_priorities_csv(text: str) -> Dict[str, List[str]]:
    out = {}
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        if len(parts) < 2:
            continue
        out[parts[0]] = [c.strip() for c in parts[1].split("|") if c.strip()]
    return out

def parse_vacations_csv(text: str, year: int, month: int) -> List[Vacation]:
    out = []
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        out.append(Vacation(parts[0], parse_date(parts[1], year, month), parse_date(parts[2], year, month)))
    return out

def parse_cabins(text: str) -> List[str]:
    return [x for x in re.split(r"[,\s]+", (text or "").strip()) if x]

def parse_holidays(text: str, year: int, month: int) -> Set[dt.date]:
    out = set()
    for tok in re.split(r"[,\s]+", (text or "").strip()):
        tok = tok.strip()
        if not tok:
            continue
        out.add(parse_date(tok, year, month))
    return out

def parse_yes_list(text: str) -> Set[str]:
    out = set()
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        out.add(line.split(",")[0].strip())
    return out

def parse_shift_pref_csv(text: str) -> Dict[str, Optional[str]]:
    out = {}
    for raw in (text or "").strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",", 1)]
        name = parts[0]
        pref = (parts[1].strip().lower() if len(parts) > 1 else "")
        if pref in ("у", "утро", "morning"):
            out[name] = "у"
        elif pref in ("в", "вечер", "evening"):
            out[name] = "в"
        elif pref in ("", "нет", "none", "-"):
            out[name] = None
        else:
            raise ValueError(f"Bad shift preference '{pref}' for '{name}'. Use у/в/нет.")
    return out

def build_vac_map(vacs: List[Vacation]) -> Dict[str, Set[dt.date]]:
    m = defaultdict(set)
    for v in vacs:
        for day in daterange(v.start, v.end):
            m[v.name].add(day)
    return m


# ----------------------------
# DB
# ----------------------------

class PhysicianDB:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(self.path)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.init_schema()

    def init_schema(self):
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS departments(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS physicians(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                department_id INTEGER,
                fte REAL NOT NULL DEFAULT 1.0,
                shift_pref TEXT,
                FOREIGN KEY(department_id) REFERENCES departments(id) ON DELETE SET NULL
            )
        """)
        self.conn.commit()

    def list_departments(self) -> List[Tuple[int, str]]:
        return list(self.conn.execute("SELECT id, name FROM departments ORDER BY name").fetchall())

    def add_department(self, name: str):
        self.conn.execute("INSERT INTO departments(name) VALUES (?)", (name.strip(),))
        self.conn.commit()

    def delete_department(self, dep_id: int):
        self.conn.execute("DELETE FROM departments WHERE id = ?", (dep_id,))
        self.conn.commit()

    def list_physicians(self) -> List[Tuple[int, str, Optional[int], str, float, Optional[str]]]:
        return list(self.conn.execute("""
            SELECT p.id, p.name, p.department_id, d.name, p.fte, p.shift_pref
            FROM physicians p
            LEFT JOIN departments d ON d.id = p.department_id
            ORDER BY p.name
        """).fetchall())

    def add_or_update_physician(self, name: str, department_id: Optional[int], fte: float, shift_pref: Optional[str]):
        row = self.conn.execute("SELECT id FROM physicians WHERE name = ?", (name.strip(),)).fetchone()
        if row:
            self.conn.execute(
                "UPDATE physicians SET department_id=?, fte=?, shift_pref=? WHERE id=?",
                (department_id, fte, shift_pref, row[0])
            )
        else:
            self.conn.execute(
                "INSERT INTO physicians(name, department_id, fte, shift_pref) VALUES (?,?,?,?)",
                (name.strip(), department_id, fte, shift_pref)
            )
        self.conn.commit()

    def delete_physician(self, phys_id: int):
        self.conn.execute("DELETE FROM physicians WHERE id = ?", (phys_id,))
        self.conn.commit()


# ----------------------------
# Solver
# ----------------------------

def required_norm(doc: Doctor, workdays: List[dt.date], vac_days: Set[dt.date]) -> int:
    base = 22 if doc.fte >= 0.99 else 11
    vac_wd = sum(1 for d in workdays if d in vac_days)
    return max(0, base - vac_wd)

def build_slots(days: List[dt.date], cabins: List[str], holidays: Set[dt.date]):
    slots = []
    for di, day in enumerate(days):
        shifts = ['р'] if (is_weekend(day) or day in holidays) else ['у', 'в']
        for sh in shifts:
            for cab in cabins:
                slots.append((di, day, sh, cab))
    return slots

def apply_priority_collision_rule(doctors: List[Doctor]) -> Tuple[List[Doctor], List[Tuple[str, str]]]:
    top_map = defaultdict(list)
    for d in doctors:
        if d.priorities:
            top_map[d.priorities[0]].append(d)
    removed = []
    for cab, ds in top_map.items():
        if len(ds) >= 3:
            half = [d for d in ds if d.fte < 0.99]
            pick = sorted(half if half else ds, key=lambda x: x.name)[-1]
            removed.append((pick.name, cab))
            pick.priorities = []
    return doctors, removed

def solve_with_cpsat(
    doctors: List[Doctor],
    vacations: List[Vacation],
    cabins: List[str],
    year: int,
    month: int,
    holidays: Set[dt.date],
    extra_ok: Set[str],
    shift_pref: Optional[Dict[str, Optional[str]]] = None,
    pref_weight: int = 3,
    enforce_morning_evening_mix: bool = True,
    extra_max: int = 6,
    time_limit_s: int = 20,
    free_label: str = "свободно",
):
    days = all_days_in_month(year, month)
    workdays = working_days_in_month(year, month)
    vac_map = build_vac_map(vacations)

    norm = {d.name: required_norm(d, workdays, vac_map.get(d.name, set())) for d in doctors}
    doctors, removed = apply_priority_collision_rule(doctors)

    doc_names = [d.name for d in doctors]
    slots = build_slots(days, cabins, holidays)
    slot_idx = {(di, sh, cab): si for si, (di, _, sh, cab) in enumerate(slots)}

    model = cp_model.CpModel()

    x = {}
    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, name in enumerate(doc_names):
            if day in vac_map.get(name, set()):
                continue
            x[(si, dj)] = model.NewBoolVar(f"x_{si}_{dj}")

    free = [model.NewBoolVar(f"free_{si}") for si in range(len(slots))]

    for si in range(len(slots)):
        vars_in = [x[(si, dj)] for dj in range(len(doc_names)) if (si, dj) in x]
        model.Add(sum(vars_in) + free[si] == 1)

    work = {}
    for di, day in enumerate(days):
        shifts = ['р'] if (is_weekend(day) or day in holidays) else ['у', 'в']
        for dj in range(len(doc_names)):
            w = model.NewBoolVar(f"work_{di}_{dj}")
            work[(di, dj)] = w
            vars_in = []
            for cab in cabins:
                for sh in shifts:
                    si = slot_idx[(di, sh, cab)]
                    if (si, dj) in x:
                        vars_in.append(x[(si, dj)])
            if vars_in:
                model.Add(sum(vars_in) == w)
            else:
                model.Add(w == 0)

    # Max 5 consecutive workdays
    for dj in range(len(doc_names)):
        for start in range(0, len(days) - 6 + 1):
            model.Add(sum(work[(di, dj)] for di in range(start, start + 6)) <= 5)

    # Norms / extra shifts
    for dj, name in enumerate(doc_names):
        total_work = sum(work[(di, dj)] for di in range(len(days)))
        if name in extra_ok:
            model.Add(total_work >= norm[name])
            model.Add(total_work <= norm[name] + extra_max)
        else:
            model.Add(total_work == norm[name])

    # By default vary morning/evening unless preference:
    # if physician has no explicit preference and has >=2 weekday shifts, require at least one morning and one evening.
    if enforce_morning_evening_mix:
        shift_pref = shift_pref or {}
        for dj, name in enumerate(doc_names):
            if shift_pref.get(name) in ('у', 'в'):
                continue
            m_vars, e_vars = [], []
            for di, day in enumerate(days):
                if is_weekend(day) or day in holidays:
                    continue
                for cab in cabins:
                    si_m = slot_idx[(di, 'у', cab)]
                    si_e = slot_idx[(di, 'в', cab)]
                    if (si_m, dj) in x:
                        m_vars.append(x[(si_m, dj)])
                    if (si_e, dj) in x:
                        e_vars.append(x[(si_e, dj)])
            if not m_vars and not e_vars:
                continue
            m_cnt = sum(m_vars) if m_vars else 0
            e_cnt = sum(e_vars) if e_vars else 0
            tot = m_cnt + e_cnt
            has_two = model.NewBoolVar(f"has_two_{dj}")
            model.Add(tot >= 2).OnlyEnforceIf(has_two)
            model.Add(tot <= 1).OnlyEnforceIf(has_two.Not())
            model.Add(m_cnt >= 1).OnlyEnforceIf(has_two)
            model.Add(e_cnt >= 1).OnlyEnforceIf(has_two)

    # Objective: fill slots, reward priorities, reward preferred shift
    obj_terms = []
    for si in range(len(slots)):
        obj_terms.append(20 * (1 - free[si]))

    for si, (di, day, sh, cab) in enumerate(slots):
        for dj, doc in enumerate(doctors):
            if (si, dj) not in x:
                continue
            if cab in doc.priorities:
                obj_terms.append((3 if doc.fte < 0.99 else 8) * x[(si, dj)])

    shift_pref = shift_pref or {}
    for dj, name in enumerate(doc_names):
        pref = shift_pref.get(name)
        if pref not in ('у', 'в'):
            continue
        for di, day in enumerate(days):
            if is_weekend(day) or day in holidays:
                continue
            pref_vars = []
            for cab in cabins:
                si = slot_idx[(di, pref, cab)]
                if (si, dj) in x:
                    pref_vars.append(x[(si, dj)])
            if pref_vars:
                obj_terms.append(int(pref_weight) * sum(pref_vars))

    model.Maximize(sum(obj_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("CP-SAT did not find a feasible solution.")

    sched = {name: {} for name in doc_names}
    for name in doc_names:
        for day in days:
            sched[name][day] = ("от", "") if day in vac_map.get(name, set()) else ("-", "")

    slot_assign = {day: [] for day in days}
    for si, (di, day, sh, cab) in enumerate(slots):
        who = free_label
        for dj, name in enumerate(doc_names):
            if (si, dj) in x and solver.Value(x[(si, dj)]) == 1:
                who = name
                break
        slot_assign[day].append((sh, cab, who))
        if who != free_label:
            sched[who][day] = (sh, cab)

    deviation = {}
    for name in doc_names:
        fact = sum(1 for day in days if sched[name][day][0] in ('у', 'в', 'р'))
        deviation[name] = fact - norm[name]

    meta = {
        "status": "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE",
        "objective": solver.ObjectiveValue(),
        "removed_priorities": removed,
    }
    return days, norm, sched, slot_assign, deviation, meta, doctors


# ----------------------------
# XLSX export
# ----------------------------

def export_xlsx_bytes(doctors, days, sched, norm, slot_assign, cabins, free_label="свободно"):
    wb = Workbook()
    ws = wb.active
    ws.title = "График врачей"

    fill_morning = PatternFill("solid", fgColor="A7D8FF")
    fill_evening = PatternFill("solid", fgColor="FFB6C1")
    fill_weekend = PatternFill("solid", fgColor="A7F3A7")
    fill_vac = PatternFill("solid", fgColor="C0C0C0")
    fill_off = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Врач").font = header_font
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=d.day)
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 11

    last_col = 2 + len(days)
    ws.cell(row=1, column=last_col, value="Смен").font = header_font
    ws.cell(row=1, column=last_col + 1, value="Комментарий").font = header_font
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 26

    for r, doc in enumerate(doctors, start=2):
        ws.cell(row=r, column=1, value=doc.name).border = border
        shift_count = 0
        for i, day in enumerate(days, start=2):
            code, cab = sched[doc.name][day]
            txt = f"{code} ({cab})" if code in ('у', 'в', 'р') else code
            if code in ('у', 'в', 'р'):
                shift_count += 1
            cell = ws.cell(row=r, column=i, value=txt)
            cell.alignment = center
            cell.border = border
            if code == 'у':
                cell.fill = fill_morning
            elif code == 'в':
                cell.fill = fill_evening
            elif code == 'р':
                cell.fill = fill_weekend
            elif code == 'от':
                cell.fill = fill_vac
            else:
                cell.fill = fill_off
        ws.cell(row=r, column=last_col, value=shift_count).border = border
        ws.cell(row=r, column=last_col + 1, value=f"ставка {doc.fte}, норма {norm[doc.name]}").border = border

    # summary
    ws2 = wb.create_sheet("Сводка по врачам")
    ws2.append(["Врач", "Ставка", "Норма", "Факт", "Откл.", "Приоритеты"])
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ('у', 'в', 'р'))
        ws2.append([doc.name, doc.fte, norm[doc.name], fact, fact - norm[doc.name], ", ".join(doc.priorities)])

    # load
    ws3 = wb.create_sheet("Загрузка кабинетов")
    ws3.append(["Дата", "День", "Смена", "Кабинет", "Врач"])
    day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for day in days:
        for sh, cab, who in slot_assign[day]:
            ws3.append([day.isoformat(), day_name[day.weekday()], sh, cab, who])

    # stats
    ws4 = wb.create_sheet("Общая статистика")
    total_slots = sum(len(slot_assign[d]) for d in days)
    free_slots = sum(1 for d in days for _, _, who in slot_assign[d] if who == free_label)
    ws4.append(["Показатель", "Значение"])
    ws4.append(["Кабинетов", len(cabins)])
    ws4.append(["Всего слотов", total_slots])
    ws4.append(["Свободно", free_slots])
    ws4.append(["Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0])

    # deviations
    ws5 = wb.create_sheet("Обоснование отклонений")
    ws5.append(["Врач", "Норма", "Факт", "Отклонение", "Пояснение"])
    for doc in doctors:
        fact = sum(1 for d in days if sched[doc.name][d][0] in ('у', 'в', 'р'))
        dev = fact - norm[doc.name]
        expl = "норма выполнена" if dev == 0 else (f"доп. смены: {dev}" if dev > 0 else f"не добрали: {-dev}")
        ws5.append([doc.name, norm[doc.name], fact, dev, expl])

    # cabins per doctor
    ws6 = wb.create_sheet("Кабинетов на врача")
    ws6.append(["Врач", "Кабинет", "Кол-во смен"])
    for doc in doctors:
        cnt = Counter()
        for d in days:
            code, cab = sched[doc.name][d]
            if code in ('у', 'в', 'р'):
                cnt[cab] += 1
        for cab, n in cnt.most_common():
            ws6.append([doc.name, cab, n])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# UI
# ----------------------------

class SchedulerTkApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Scheduler — OR-Tools + DB")
        self.root.geometry("1600x920")
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.db = PhysicianDB(os.path.join(os.path.dirname(__file__), "physicians.db"))
        self.last_result = None

        self.var_year = tk.IntVar(value=2025)
        self.var_month = tk.IntVar(value=10)
        self.var_time = tk.IntVar(value=20)
        self.var_extra_max = tk.IntVar(value=6)
        self.var_pref_weight = tk.IntVar(value=3)
        self.var_enforce_mix = tk.BooleanVar(value=True)

        self._build_ui()
        self._fill_defaults()
        self.refresh_departments()
        self.refresh_physicians()

    def _scrolled_text(self, parent, height=8):
        frame = ttk.Frame(parent)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        txt = tk.Text(frame, height=height, wrap="none")
        sy = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        sx = ttk.Scrollbar(frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        txt.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        return frame, txt

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=8)
        main.grid(row=0, column=0, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        # top controls
        top = ttk.Frame(main)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(top, text="Год").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(top, from_=2020, to=2035, textvariable=self.var_year, width=6).grid(row=0, column=1, padx=4)
        ttk.Label(top, text="Месяц").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(top, from_=1, to=12, textvariable=self.var_month, width=4).grid(row=0, column=3, padx=4)
        ttk.Label(top, text="Time limit").grid(row=0, column=4, sticky="w")
        ttk.Spinbox(top, from_=3, to=120, textvariable=self.var_time, width=5).grid(row=0, column=5, padx=4)
        ttk.Label(top, text="Extra max").grid(row=0, column=6, sticky="w")
        ttk.Spinbox(top, from_=0, to=20, textvariable=self.var_extra_max, width=5).grid(row=0, column=7, padx=4)
        ttk.Label(top, text="Вес pref смены").grid(row=0, column=8, sticky="w")
        ttk.Spinbox(top, from_=0, to=10, textvariable=self.var_pref_weight, width=5).grid(row=0, column=9, padx=4)
        ttk.Checkbutton(top, text="Варьировать утро/вечер по умолчанию", variable=self.var_enforce_mix).grid(row=0, column=10, padx=8)
        ttk.Button(top, text="Recompute", command=self.on_recompute).grid(row=0, column=11, padx=8)
        ttk.Button(top, text="Export XLSX", command=self.on_export).grid(row=0, column=12, padx=4)

        nb = ttk.Notebook(main)
        nb.grid(row=1, column=0, sticky="nsew")
        self.nb = nb

        self.tab_input = ttk.Frame(nb)
        self.tab_output = ttk.Frame(nb)
        nb.add(self.tab_input, text="Ввод")
        nb.add(self.tab_output, text="Результат")

        self._build_input_tab()
        self._build_output_tab()

        self.status = tk.StringVar(value="Готово.")
        ttk.Label(main, textvariable=self.status, anchor="w").grid(row=2, column=0, sticky="ew", pady=(6, 0))

    def _build_input_tab(self):
        t = self.tab_input
        t.rowconfigure(0, weight=1)
        t.columnconfigure(0, weight=1)

        pw = ttk.Panedwindow(t, orient="horizontal")
        pw.grid(row=0, column=0, sticky="nsew")

        left = ttk.Frame(pw, padding=8)
        right = ttk.Frame(pw, padding=8)
        pw.add(left, weight=1)
        pw.add(right, weight=3)

        # Left: database pane
        left.rowconfigure(3, weight=1)
        left.columnconfigure(0, weight=1)

        ttk.Label(left, text="База данных").grid(row=0, column=0, sticky="w")

        dep_frame = ttk.LabelFrame(left, text="Отделения", padding=6)
        dep_frame.grid(row=1, column=0, sticky="nsew", pady=(6, 6))
        dep_frame.rowconfigure(1, weight=1)
        dep_frame.columnconfigure(0, weight=1)

        dep_top = ttk.Frame(dep_frame)
        dep_top.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        self.ent_dep = ttk.Entry(dep_top)
        self.ent_dep.grid(row=0, column=0, sticky="ew")
        dep_top.columnconfigure(0, weight=1)
        ttk.Button(dep_top, text="Add", command=self.on_add_department).grid(row=0, column=1, padx=4)
        ttk.Button(dep_top, text="Delete", command=self.on_delete_department).grid(row=0, column=2, padx=4)

        self.lst_dep = tk.Listbox(dep_frame, exportselection=False, height=8)
        self.lst_dep.grid(row=1, column=0, sticky="nsew")
        self.lst_dep.bind("<<ListboxSelect>>", lambda e: self.refresh_physicians())

        phys_frame = ttk.LabelFrame(left, text="Врачи", padding=6)
        phys_frame.grid(row=2, column=0, sticky="nsew", pady=(6, 6))
        phys_frame.rowconfigure(2, weight=1)
        phys_frame.columnconfigure(0, weight=1)

        form = ttk.Frame(phys_frame)
        form.grid(row=0, column=0, sticky="ew")
        ttk.Label(form, text="ФИО").grid(row=0, column=0, sticky="w")
        self.ent_phys_name = ttk.Entry(form)
        self.ent_phys_name.grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Label(form, text="FTE").grid(row=1, column=0, sticky="w")
        self.cmb_phys_fte = ttk.Combobox(form, values=["1.0", "0.5"], width=8)
        self.cmb_phys_fte.set("1.0")
        self.cmb_phys_fte.grid(row=1, column=1, sticky="w", padx=4)
        ttk.Label(form, text="Pref shift").grid(row=2, column=0, sticky="w")
        self.cmb_phys_pref = ttk.Combobox(form, values=["", "у", "в"], width=8)
        self.cmb_phys_pref.grid(row=2, column=1, sticky="w", padx=4)
        form.columnconfigure(1, weight=1)

        btns = ttk.Frame(phys_frame)
        btns.grid(row=1, column=0, sticky="ew", pady=4)
        ttk.Button(btns, text="Add / Update", command=self.on_add_update_physician).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Delete", command=self.on_delete_physician).grid(row=0, column=1, padx=4)
        ttk.Button(btns, text="Load DB → inputs", command=self.on_load_db_to_inputs).grid(row=0, column=2, padx=4)

        self.tree_phys_db = ttk.Treeview(phys_frame, columns=("name", "dept", "fte", "pref"), show="headings", height=12)
        for c, w in [("name", 200), ("dept", 120), ("fte", 50), ("pref", 60)]:
            self.tree_phys_db.heading(c, text=c)
            self.tree_phys_db.column(c, width=w, anchor="w")
        self.tree_phys_db.grid(row=2, column=0, sticky="nsew")
        self.tree_phys_db.bind("<<TreeviewSelect>>", self.on_select_physician)

        # Right: editable inputs
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        nb_in = ttk.Notebook(right)
        nb_in.grid(row=0, column=0, sticky="nsew")

        tabs = []
        for title in ["Врачи", "Приоритеты", "Отпуска", "Кабинеты", "Праздники", "Доп. смены", "Pref смены"]:
            fr = ttk.Frame(nb_in)
            nb_in.add(fr, text=title)
            tabs.append(fr)

        _, self.txt_doctors = self._scrolled_text(tabs[0], 14)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[0].rowconfigure(0, weight=1); tabs[0].columnconfigure(0, weight=1)

        _, self.txt_prio = self._scrolled_text(tabs[1], 14)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[1].rowconfigure(0, weight=1); tabs[1].columnconfigure(0, weight=1)

        _, self.txt_vac = self._scrolled_text(tabs[2], 14)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[2].rowconfigure(0, weight=1); tabs[2].columnconfigure(0, weight=1)

        _, self.txt_cabins = self._scrolled_text(tabs[3], 8)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[3].rowconfigure(0, weight=1); tabs[3].columnconfigure(0, weight=1)

        _, self.txt_holidays = self._scrolled_text(tabs[4], 8)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[4].rowconfigure(0, weight=1); tabs[4].columnconfigure(0, weight=1)

        _, self.txt_extra_ok = self._scrolled_text(tabs[5], 8)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[5].rowconfigure(0, weight=1); tabs[5].columnconfigure(0, weight=1)

        _, self.txt_shift_pref = self._scrolled_text(tabs[6], 8)
        _.grid(row=0, column=0, sticky="nsew")
        tabs[6].rowconfigure(0, weight=1); tabs[6].columnconfigure(0, weight=1)

    def _build_output_tab(self):
        t = self.tab_output
        t.rowconfigure(1, weight=1)
        t.columnconfigure(0, weight=1)

        info = ttk.Frame(t, padding=8)
        info.grid(row=0, column=0, sticky="ew")
        info.columnconfigure(1, weight=1)
        ttk.Label(info, text="Status").grid(row=0, column=0, sticky="w")
        self.lbl_solution = ttk.Label(info, text="—")
        self.lbl_solution.grid(row=0, column=1, sticky="w")
        ttk.Label(info, text="Removed priority").grid(row=1, column=0, sticky="w")
        self.lbl_removed = ttk.Label(info, text="—")
        self.lbl_removed.grid(row=1, column=1, sticky="w")

        nb = ttk.Notebook(t)
        nb.grid(row=1, column=0, sticky="nsew", padx=8, pady=8)
        self.nb_out = nb

        self.tab_out_grid = ttk.Frame(nb)
        self.tab_out_summary = ttk.Frame(nb)
        self.tab_out_load = ttk.Frame(nb)
        self.tab_out_stats = ttk.Frame(nb)
        self.tab_out_dev = ttk.Frame(nb)
        self.tab_out_cabdoc = ttk.Frame(nb)

        nb.add(self.tab_out_grid, text="График врачей")
        nb.add(self.tab_out_summary, text="Сводка")
        nb.add(self.tab_out_load, text="Загрузка кабинетов")
        nb.add(self.tab_out_stats, text="Статистика")
        nb.add(self.tab_out_dev, text="Отклонения")
        nb.add(self.tab_out_cabdoc, text="Кабинеты на врача")

        self._build_grid_view(self.tab_out_grid)
        self.tree_summary = self._make_tree(self.tab_out_summary, ("Врач", "Ставка", "Норма", "Факт", "Откл.", "Приоритеты"))
        self.tree_load = self._make_tree(self.tab_out_load, ("Дата", "День", "Смена", "Кабинет", "Врач"))
        self.tree_stats = self._make_tree(self.tab_out_stats, ("Показатель", "Значение"))
        self.tree_dev = self._make_tree(self.tab_out_dev, ("Врач", "Норма", "Факт", "Отклонение", "Пояснение"))
        self.tree_cabdoc = self._make_tree(self.tab_out_cabdoc, ("Врач", "Кабинет", "Кол-во смен"))

    def _build_grid_view(self, parent):
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        if HAS_TKSHEET:
            self.sheet = Sheet(parent)
            self.sheet.enable_bindings(("single_select", "row_select", "column_select", "drag_select",
                                        "column_width_resize", "double_click_column_resize",
                                        "arrowkeys", "right_click_popup_menu", "rc_select",
                                        "copy", "edit_cell"))
            self.sheet.grid(row=0, column=0, sticky="nsew")
            self.grid_widget = "tksheet"
        else:
            frame = ttk.Frame(parent)
            frame.grid(row=0, column=0, sticky="nsew")
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)
            self.tree = ttk.Treeview(frame, show="headings")
            sy = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
            sx = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
            self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
            self.tree.grid(row=0, column=0, sticky="nsew")
            sy.grid(row=0, column=1, sticky="ns")
            sx.grid(row=1, column=0, sticky="ew")
            self.grid_widget = "treeview"

    def _make_tree(self, parent, columns):
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        frame = ttk.Frame(parent)
        frame.grid(row=0, column=0, sticky="nsew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        tree = ttk.Treeview(frame, show="headings", columns=columns)
        sy = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        sx = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=140 if c != "Врач" else 280, anchor="w")
        return tree

    def _fill_defaults(self):
        self._set_text(self.txt_doctors, """Боброва Эльвира Анатольевна,1
Буданова Татьяна Владимировна,0.5
Варвус Иван Михайлович,1
Гагкаева Мария Аркадьевна,1
Гущина Юлия Викторовна,1
Джафаров Башир Темирханович,0.5
Догузова Александра Гочевна,1
Евдокимова Жанна Владимировна,1
Егуренкова Виктория Сергеевна,1
Захаров Александр Сергеевич,1
Исаев Абсалам Темурланович,1
Китков Игорь Игоревич,1
Кокорина Серафима Евгеньевна,1
Королькова Александра Викторовна,0.5
Короткова Полина Александровна,1
Лукьяненко Владимир Александрович,1
Маркин Александр Андреевич,1
Марфутов Василий Васильевич,1
Медведева Мария Вячеславовна,1
Найданова Лидия Вадимовна,1
Нам Ирина Николаевна,1
Петрова Полина Юрьевна,1
Пономарева Юлия Константиновна,1
Прокофьев Евгений Андреевич,1
Пусь Юлия Владимировна,1
Танкиева Фатима Умат-Гиреевна,1
Тетерев Ярослав Тарасович,1
Федорова Оксана Сергеевна,1
Харитонова Наталья Игоревна,1
Шепель Артем Олегович,1
Юдин Алексей Александрович,1
Юдина Виктория Дмитриевна,1""")

        self._set_text(self.txt_prio, """Пономарева Юлия Константиновна,2А03
Лукьяненко Владимир Александрович,2А03
Королькова Александра Викторовна,2А03
Харитонова Наталья Игоревна,2А04
Исаев Абсалам Темурланович,2А04
Короткова Полина Александровна,2А04
Гущина Юлия Викторовна,2А05
Федорова Оксана Сергеевна,2А05
Кокорина Серафима Евгеньевна,2А05
Егуренкова Виктория Сергеевна,2А06
Евдокимова Жанна Владимировна,2А06
Буданова Татьяна Владимировна,2А06
Боброва Эльвира Анатольевна,2А07
Марфутов Василий Васильевич,2А07
Догузова Александра Гочевна,2А07
Медведева Мария Вячеславовна,2А08
Юдин Алексей Александрович,2А08
Юдина Виктория Дмитриевна,2А33
Шепель Артем Олегович,2А33
Китков Игорь Игоревич,2А34
Гагкаева Мария Аркадьевна,2А34
Варвус Иван Михайлович,2А49
Маркин Александр Андреевич,2А49
Тетерев Ярослав Тарасович,2А52
Джафаров Башир Темирханович,2А52
Петрова Полина Юрьевна,2А52
Пусь Юлия Владимировна,2А54
Прокофьев Евгений Андреевич,2А54
Захаров Александр Сергеевич,2А54
Найданова Лидия Вадимовна,2А55
Нам Ирина Николаевна,2А55|2А54
Танкиева Фатима Умат-Гиреевна,2А55""")

        self._set_text(self.txt_vac, """Боброва Эльвира Анатольевна,01.10.2025,12.10.2025
Тетерев Ярослав Тарасович,29.09.2025,12.10.2025
Исаев Абсалам Темурланович,06.10.2025,19.10.2025
Гагкаева Мария Аркадьевна,01.10.2025,14.10.2025
Гагкаева Мария Аркадьевна,20.10.2025,20.10.2025
Гагкаева Мария Аркадьевна,27.10.2025,27.10.2025
Егуренкова Виктория Сергеевна,06.10.2025,19.10.2025""")

        self._set_text(self.txt_cabins, "2А03, 2А04, 2А05, 2А06, 2А07, 2А08, 2А33, 2А34, 2А49, 2А52, 2А54, 2А55")
        self._set_text(self.txt_holidays, "")
        self._set_text(self.txt_extra_ok, "")
        self._set_text(self.txt_shift_pref, "# ФИО,у/в/нет")

    def _get_text(self, txt):
        return txt.get("1.0", "end").strip()

    def _set_text(self, txt, value):
        txt.delete("1.0", "end")
        txt.insert("1.0", value)

    # DB handlers
    def refresh_departments(self):
        self.lst_dep.delete(0, "end")
        self.dep_rows = self.db.list_departments()
        for _id, name in self.dep_rows:
            self.lst_dep.insert("end", name)

    def selected_department_id(self):
        sel = self.lst_dep.curselection()
        if not sel:
            return None
        return self.dep_rows[sel[0]][0]

    def refresh_physicians(self):
        for it in self.tree_phys_db.get_children():
            self.tree_phys_db.delete(it)
        self.phys_rows = self.db.list_physicians()
        dep_id = self.selected_department_id()
        for row in self.phys_rows:
            pid, name, pdep_id, dep_name, fte, shift_pref = row
            if dep_id is not None and pdep_id != dep_id:
                continue
            self.tree_phys_db.insert("", "end", iid=str(pid), values=(name, dep_name or "", fte, shift_pref or ""))

    def on_add_department(self):
        name = self.ent_dep.get().strip()
        if not name:
            return
        try:
            self.db.add_department(name)
            self.ent_dep.delete(0, "end")
            self.refresh_departments()
        except Exception as e:
            messagebox.showerror("DB error", str(e))

    def on_delete_department(self):
        dep_id = self.selected_department_id()
        if dep_id is None:
            return
        if messagebox.askyesno("Delete", "Delete selected department?"):
            self.db.delete_department(dep_id)
            self.refresh_departments()
            self.refresh_physicians()

    def on_add_update_physician(self):
        name = self.ent_phys_name.get().strip()
        if not name:
            return
        dep_id = self.selected_department_id()
        try:
            fte = float(self.cmb_phys_fte.get().strip() or "1.0")
        except ValueError:
            messagebox.showerror("Input error", "FTE must be 1.0 or 0.5")
            return
        shift_pref = self.cmb_phys_pref.get().strip() or None
        try:
            self.db.add_or_update_physician(name, dep_id, fte, shift_pref)
            self.refresh_physicians()
        except Exception as e:
            messagebox.showerror("DB error", str(e))

    def on_delete_physician(self):
        sel = self.tree_phys_db.selection()
        if not sel:
            return
        pid = int(sel[0])
        if messagebox.askyesno("Delete", "Delete selected physician?"):
            self.db.delete_physician(pid)
            self.refresh_physicians()

    def on_select_physician(self, event=None):
        sel = self.tree_phys_db.selection()
        if not sel:
            return
        pid = int(sel[0])
        row = next((r for r in self.phys_rows if r[0] == pid), None)
        if not row:
            return
        _, name, _, _, fte, pref = row
        self.ent_phys_name.delete(0, "end")
        self.ent_phys_name.insert(0, name)
        self.cmb_phys_fte.set(str(fte))
        self.cmb_phys_pref.set(pref or "")

    def on_load_db_to_inputs(self):
        rows = self.db.list_physicians()
        doctors_lines, pref_lines = [], []
        for _, name, _, _, fte, shift_pref in rows:
            doctors_lines.append(f"{name},{fte}")
            if shift_pref:
                pref_lines.append(f"{name},{shift_pref}")
        self._set_text(self.txt_doctors, "\n".join(doctors_lines))
        current = self._get_text(self.txt_shift_pref)
        merged = "\n".join([ln for ln in [current, *pref_lines] if ln]).strip()
        self._set_text(self.txt_shift_pref, merged)
        messagebox.showinfo("Loaded", "Physicians from DB loaded into input panes.")

    # Actions
    def on_recompute(self):
        try:
            self.status.set("Solving CP-SAT...")
            self.root.update_idletasks()

            year = int(self.var_year.get())
            month = int(self.var_month.get())
            time_limit = int(self.var_time.get())
            extra_max = int(self.var_extra_max.get())
            pref_weight = int(self.var_pref_weight.get())
            enforce_mix = bool(self.var_enforce_mix.get())

            doctors = parse_doctors_csv(self._get_text(self.txt_doctors))
            pr_map = parse_priorities_csv(self._get_text(self.txt_prio))
            for d in doctors:
                d.priorities = pr_map.get(d.name, [])

            vacations = parse_vacations_csv(self._get_text(self.txt_vac), year, month)
            cabins = parse_cabins(self._get_text(self.txt_cabins))
            holidays = parse_holidays(self._get_text(self.txt_holidays), year, month)
            extra_ok = parse_yes_list(self._get_text(self.txt_extra_ok))
            shift_pref = parse_shift_pref_csv(self._get_text(self.txt_shift_pref))

            # Merge DB physicians shift preferences as defaults if text not specified
            for _, name, _, _, _fte, db_pref in self.db.list_physicians():
                if name not in shift_pref and db_pref:
                    shift_pref[name] = db_pref

            if not doctors:
                raise ValueError("Doctors list is empty.")
            if not cabins:
                raise ValueError("Cabins list is empty.")

            days, norm, sched, slot_assign, deviation, meta, doctors_final = solve_with_cpsat(
                doctors=doctors,
                vacations=vacations,
                cabins=cabins,
                year=year,
                month=month,
                holidays=holidays,
                extra_ok=extra_ok,
                shift_pref=shift_pref,
                pref_weight=pref_weight,
                enforce_morning_evening_mix=enforce_mix,
                extra_max=extra_max,
                time_limit_s=time_limit,
            )

            self.last_result = (days, norm, sched, slot_assign, deviation, meta, doctors_final, cabins)
            self._render_all()
            self.status.set("Ready.")
        except Exception as e:
            self.status.set("Error.")
            messagebox.showerror("Error", str(e))

    def on_export(self):
        if not self.last_result:
            messagebox.showwarning("No data", "Run Recompute first.")
            return
        days, norm, sched, slot_assign, deviation, meta, doctors_final, cabins = self.last_result
        path = filedialog.asksaveasfilename(
            title="Save XLSX",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"schedule_{self.var_year.get()}_{self.var_month.get():02d}.xlsx"
        )
        if not path:
            return
        data = export_xlsx_bytes(doctors_final, days, sched, norm, slot_assign, cabins)
        with open(path, "wb") as f:
            f.write(data)
        messagebox.showinfo("Saved", path)

    def _render_all(self):
        days, norm, sched, slot_assign, deviation, meta, doctors, cabins = self.last_result

        # grid sheet
        cols = ["Врач"] + [str(d.day) for d in days] + ["Смен", "Комментарий"]
        rows = []
        for doc in doctors:
            shift_count = 0
            row = [doc.name]
            for day in days:
                code, cab = sched[doc.name][day]
                if code in ('у', 'в', 'р'):
                    row.append(f"{code} ({cab})")
                    shift_count += 1
                else:
                    row.append(code)
            row.append(str(shift_count))
            row.append(f"ставка {doc.fte}, норма {norm[doc.name]}")
            rows.append(row)

        if self.grid_widget == "tksheet":
            self.sheet.set_sheet_data(rows)
            self.sheet.headers(cols)
            widths = [260] + [90] * len(days) + [70, 220]
            try:
                self.sheet.set_column_widths(widths)
            except Exception:
                pass
            try:
                self.sheet.dehighlight_all()
                self.sheet.delete_all_cell_options()
            except Exception:
                pass
            def bg(code):
                return {'у': '#A7D8FF', 'в': '#FFB6C1', 'р': '#A7F3A7', 'от': '#C0C0C0'}.get(code)
            for r, doc in enumerate(doctors):
                for c, day in enumerate(days, start=1):
                    code, _ = sched[doc.name][day]
                    color = bg(code)
                    if color:
                        try:
                            self.sheet.highlight_cells(row=r, column=c, bg=color, fg="black")
                        except Exception:
                            pass
            try:
                self.sheet.refresh()
            except Exception:
                pass
        else:
            self.tree["columns"] = cols
            for c in cols:
                self.tree.heading(c, text=c)
                self.tree.column(c, width=260 if c == "Врач" else (220 if c == "Комментарий" else 90), anchor="w" if c in ("Врач", "Комментарий") else "center")
            for it in self.tree.get_children():
                self.tree.delete(it)
            for row in rows:
                self.tree.insert("", "end", values=row)

        # summary
        for it in self.tree_summary.get_children():
            self.tree_summary.delete(it)
        for doc in doctors:
            fact = sum(1 for d in days if sched[doc.name][d][0] in ('у', 'в', 'р'))
            self.tree_summary.insert("", "end", values=(doc.name, doc.fte, norm[doc.name], fact, fact - norm[doc.name], ", ".join(doc.priorities)))

        # load
        for it in self.tree_load.get_children():
            self.tree_load.delete(it)
        day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        for day in days:
            for sh, cab, who in slot_assign[day]:
                self.tree_load.insert("", "end", values=(day.isoformat(), day_name[day.weekday()], sh, cab, who))

        # stats
        for it in self.tree_stats.get_children():
            self.tree_stats.delete(it)
        total_slots = sum(len(slot_assign[d]) for d in days)
        free_slots = sum(1 for d in days for _, _, who in slot_assign[d] if who == "свободно")
        for row in [
            ("Кабинетов", len(cabins)),
            ("Всего слотов", total_slots),
            ("Свободно", free_slots),
            ("Заполнение %", (total_slots - free_slots) / total_slots if total_slots else 0),
            ("Статус решения", meta.get("status", "—")),
            ("Objective", meta.get("objective", "—")),
        ]:
            self.tree_stats.insert("", "end", values=row)

        # deviation
        for it in self.tree_dev.get_children():
            self.tree_dev.delete(it)
        for doc in doctors:
            fact = sum(1 for d in days if sched[doc.name][d][0] in ('у', 'в', 'р'))
            dev = fact - norm[doc.name]
            vac_count = sum(1 for d in days if sched[doc.name][d][0] == 'от')
            expl = []
            if vac_count:
                expl.append(f"отпуск: {vac_count} дн.")
            if dev > 0:
                expl.append(f"доп. смены: {dev}")
            elif dev < 0:
                expl.append(f"не добрали: {-dev}")
            else:
                expl.append("норма выполнена")
            self.tree_dev.insert("", "end", values=(doc.name, norm[doc.name], fact, dev, "; ".join(expl)))

        # cabins per doc
        for it in self.tree_cabdoc.get_children():
            self.tree_cabdoc.delete(it)
        for doc in doctors:
            cnt = Counter()
            for d in days:
                code, cab = sched[doc.name][d]
                if code in ('у', 'в', 'р'):
                    cnt[cab] += 1
            for cab, n in cnt.most_common():
                self.tree_cabdoc.insert("", "end", values=(doc.name, cab, n))

        self.lbl_solution.config(text=meta.get("status", "—"))
        removed = meta.get("removed_priorities") or []
        self.lbl_removed.config(text="; ".join(f"{a} (снят {b})" for a, b in removed) if removed else "—")

    def run(self):
        self.root.mainloop()


def main():
    SchedulerTkApp().run()


if __name__ == "__main__":
    main()
